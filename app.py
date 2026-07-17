from flask import Flask, render_template, redirect, url_for, request, session, flash, send_file, jsonify
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime, date, timedelta, timezone
from openpyxl import Workbook, load_workbook
from functools import wraps
from werkzeug.utils import secure_filename
import sqlite3
import os
import json
import csv
import io
import unicodedata
import shutil
from urllib.parse import quote

app = Flask(__name__)
app.secret_key = "troque_essa_chave_depois"

@app.context_processor
def inject_configuracoes():
    try:
        return {
            "config_loja": obter_configuracoes_loja()
        }
    except Exception:
        return {
            "config_loja": None
        }

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, "crm_start.db")
HISTORICO_IMPORTACOES_PRODUTOS_DIR = os.path.join(
    BASE_DIR,
    "exports",
    "importacoes_produtos"
)

@app.template_filter("datetime_br")
def datetime_br(valor):
    if not valor:
        return "-"

    try:
        texto = str(valor).strip()

        try:
            data_utc = datetime.strptime(texto, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            data_utc = datetime.fromisoformat(texto)

        if data_utc.tzinfo is None:
            data_utc = data_utc.replace(tzinfo=timezone.utc)

        fuso_brasil = timezone(timedelta(hours=-3))
        data_brasil = data_utc.astimezone(fuso_brasil)

        return data_brasil.strftime("%d/%m/%Y %H:%M")

    except Exception as erro:
        print(f"Erro ao formatar data: {erro}")
        return valor
    
@app.template_filter("date_br")
def date_br(valor):
    if not valor:
        return "-"

    try:
        texto = str(valor).strip()

        try:
            data_utc = datetime.strptime(texto, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            data_utc = datetime.fromisoformat(texto)

        if data_utc.tzinfo is None:
            data_utc = data_utc.replace(tzinfo=timezone.utc)

        fuso_brasil = timezone(timedelta(hours=-3))
        data_brasil = data_utc.astimezone(fuso_brasil)

        return data_brasil.strftime("%d/%m/%Y")

    except Exception as erro:
        print(f"Erro ao formatar data: {erro}")
        return valor


@app.template_filter("moeda_br")
def moeda_br(valor):
    try:
        numero = float(valor or 0)
    except (TypeError, ValueError):
        numero = 0

    formatado = f"{numero:,.2f}"
    formatado = formatado.replace(",", "_").replace(".", ",").replace("_", ".")
    return f"R$ {formatado}"


@app.template_filter("numero_br")
def numero_br(valor, casas=1):
    try:
        numero = float(valor or 0)
        casas_decimais = max(0, min(int(casas), 4))
    except (TypeError, ValueError):
        numero = 0
        casas_decimais = 1

    return f"{numero:.{casas_decimais}f}".replace(".", ",")
    
@app.template_global("whatsapp_link")
def whatsapp_link(telefone, mensagem):
    if not telefone:
        return "#"

    numero = "".join(c for c in str(telefone) if c.isdigit())

    if numero.startswith("0"):
        numero = numero[1:]

    if not numero.startswith("55"):
        numero = "55" + numero

    mensagem_url = quote(mensagem)

    return f"https://wa.me/{numero}?text={mensagem_url}"

@app.template_global("pagination_url")
def pagination_url(endpoint, args=None, page_param="page", page=1, anchor=None):
    args = dict(args or {})
    args[page_param] = page

    url = url_for(endpoint, **args)

    if anchor:
        url += f"#{anchor}"

    return url

def obter_caixa_aberto(conn):
    return conn.execute("""
        SELECT *
        FROM caixas
        WHERE status = 'ABERTO'
        ORDER BY id DESC
        LIMIT 1
    """).fetchone()


def normalizar_forma_pagamento_caixa(forma_pagamento):
    texto = (forma_pagamento or "").strip().lower()

    if "dinheiro" in texto:
        return "dinheiro"

    if "pix" in texto:
        return "pix"

    if "cart" in texto or "credito" in texto or "crédito" in texto or "debito" in texto or "débito" in texto:
        return "cartao"

    return "outros"


def calcular_resumo_caixa(conn, caixa_id):
    pagamentos = conn.execute("""
        SELECT
            vp.forma_pagamento,
            SUM(vp.valor) AS total
        FROM venda_pagamentos vp
        JOIN vendas v ON v.id = vp.venda_id
        WHERE v.caixa_id = ?
        AND v.status = 'CONCLUIDA'
        GROUP BY vp.forma_pagamento
    """, (caixa_id,)).fetchall()

    total_dinheiro = 0
    total_pix = 0
    total_cartao = 0
    total_outros = 0

    for pagamento in pagamentos:
        forma = pagamento["forma_pagamento"]
        total = pagamento["total"] or 0

        if forma == "DINHEIRO":
            total_dinheiro += total
        elif forma == "PIX":
            total_pix += total
        elif forma == "CARTAO":
            total_cartao += total
        else:
            total_outros += total

    total_vendas = total_dinheiro + total_pix + total_cartao + total_outros

    movimentos = conn.execute("""
        SELECT
            tipo,
            COALESCE(SUM(valor), 0) AS total
        FROM caixa_movimentacoes
        WHERE caixa_id = ?
        GROUP BY tipo
    """, (caixa_id,)).fetchall()

    entradas_manuais = 0
    saidas_manuais = 0

    for movimento in movimentos:
        if movimento["tipo"] == "ENTRADA":
            entradas_manuais += float(movimento["total"] or 0)
        elif movimento["tipo"] == "SAIDA":
            saidas_manuais += float(movimento["total"] or 0)

    caixa = conn.execute("""
        SELECT valor_inicial
        FROM caixas
        WHERE id = ?
    """, (caixa_id,)).fetchone()

    valor_inicial = float(caixa["valor_inicial"] or 0) if caixa else 0

    valor_esperado = valor_inicial + total_dinheiro + entradas_manuais - saidas_manuais

    return {
        "total_vendas": total_vendas,
        "total_dinheiro": total_dinheiro,
        "total_pix": total_pix,
        "total_cartao": total_cartao,
        "total_outros": total_outros,
        "entradas_manuais": entradas_manuais,
        "saidas_manuais": saidas_manuais,
        "valor_esperado": valor_esperado
    }

def normalizar_pagina(valor):
    try:
        page = int(valor)
    except (TypeError, ValueError):
        page = 1

    return max(page, 1)


def calcular_paginacao(total_registros, page, per_page):
    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas == 0:
        total_paginas = 1

    if page > total_paginas:
        page = total_paginas

    offset = (page - 1) * per_page

    return page, total_paginas, offset


def buscar_clientes_dashboard(conn, tipo, page=1, per_page=10):
    configuracoes = {
        "reativar": {
            "where": "rv.ultima_compra IS NOT NULL AND date(rv.ultima_compra) <= date('now', '-30 days')",
            "order": "rv.ultima_compra ASC",
            "page_param": "page_reativar",
        },
        "vip": {
            "where": "COALESCE(rv.quantidade_compras, 0) > 0",
            "order": "rv.total_comprado DESC",
            "page_param": "page_vip",
        },
        "sem_compra": {
            "where": "COALESCE(rv.quantidade_compras, 0) = 0",
            "order": "c.created_at DESC",
            "page_param": "page_sem_compra",
        },
        "sem_tags": {
            "where": "(tg.tags IS NULL OR tg.tags = '')",
            "order": "c.created_at DESC",
            "page_param": "page_sem_tags",
        },
        "recentes_30": {
            "where": "rv.ultima_compra IS NOT NULL AND date(rv.ultima_compra) >= date('now', '-30 days')",
            "order": "rv.ultima_compra DESC",
            "page_param": "page_recentes_30",
        },
    }

    config = configuracoes[tipo]

    base_from = f"""
        FROM clientes c
        LEFT JOIN (
            SELECT
                cliente_id,
                MAX(data_venda) AS ultima_compra,
                COUNT(id) AS quantidade_compras,
                COALESCE(SUM(valor_total), 0) AS total_comprado
            FROM vendas
            WHERE status = 'CONCLUIDA'
              AND cliente_id IS NOT NULL
            GROUP BY cliente_id
        ) rv ON rv.cliente_id = c.id
        LEFT JOIN (
            SELECT
                cliente_id,
                GROUP_CONCAT(tag, ', ') AS tags
            FROM cliente_tags
            GROUP BY cliente_id
        ) tg ON tg.cliente_id = c.id
        WHERE c.ativo = 1
          AND {config["where"]}
    """

    total_registros = conn.execute(f"""
        SELECT COUNT(*) AS total
        {base_from}
    """).fetchone()["total"]

    page, total_paginas, offset = calcular_paginacao(total_registros, page, per_page)

    rows = conn.execute(f"""
        SELECT
            c.id,
            c.nome,
            c.telefone,
            c.created_at,
            rv.ultima_compra,
            COALESCE(rv.quantidade_compras, 0) AS quantidade_compras,
            COALESCE(rv.total_comprado, 0) AS total_comprado,
            COALESCE(tg.tags, '-') AS tags
        {base_from}
        ORDER BY {config["order"]}
        LIMIT ? OFFSET ?
    """, (per_page, offset)).fetchall()

    return {
        "rows": rows,
        "total": total_registros,
        "page": page,
        "total_pages": total_paginas,
        "page_param": config["page_param"],
    }

def buscar_produtos_dashboard(conn, tipo, page=1, per_page=10):
    configuracoes = {
        "baixo": {
            "where": "p.estoque_atual > 0 AND p.estoque_atual <= p.estoque_minimo",
            "order": "p.estoque_atual ASC, p.nome ASC",
            "page_param": "page_produtos_baixo",
        },
        "sem": {
            "where": "p.estoque_atual = 0",
            "order": "p.nome ASC",
            "page_param": "page_produtos_sem",
        },
    }

    config = configuracoes[tipo]

    total_registros = conn.execute(f"""
        SELECT COUNT(*) AS total
        FROM produtos p
        WHERE p.ativo = 1
          AND {config["where"]}
    """).fetchone()["total"]

    page, total_paginas, offset = calcular_paginacao(
        total_registros,
        page,
        per_page
    )

    rows = conn.execute(f"""
        SELECT
            p.id,
            p.nome,
            p.categoria,
            p.marca,
            p.sku,
            p.estoque_atual,
            p.estoque_minimo,
            p.preco_custo,
            p.preco_venda
        FROM produtos p
        WHERE p.ativo = 1
          AND {config["where"]}
        ORDER BY {config["order"]}
        LIMIT ? OFFSET ?
    """, (
        per_page,
        offset
    )).fetchall()

    return {
        "rows": rows,
        "total": total_registros,
        "page": page,
        "total_pages": total_paginas,
        "page_param": config["page_param"],
    }
    
def aplicar_variaveis_mensagem(mensagem, cliente, config, resumo=None, tags=None):
    if resumo is None:
        resumo = {}

    if tags is None:
        tags = []

    tags_texto = ", ".join(tags) if tags else "-"

    ultima_compra = resumo["ultima_compra"] if resumo and resumo["ultima_compra"] else "-"
    total_comprado = resumo["total_comprado"] if resumo and resumo["total_comprado"] else 0

    variaveis = {
        "{cliente_nome}": cliente["nome"] or "",
        "{cliente_telefone}": cliente["telefone"] or "",
        "{loja_nome}": config["nome_loja"] if config and config["nome_loja"] else "Start Eletrônicos",
        "{loja_telefone}": config["telefone"] if config and config["telefone"] else "",
        "{loja_instagram}": config["instagram"] if config and config["instagram"] else "",
        "{ultima_compra}": str(ultima_compra),
        "{total_comprado}": f"R$ {float(total_comprado):.2f}",
        "{tags}": tags_texto
    }

    mensagem_final = mensagem

    for chave, valor in variaveis.items():
        mensagem_final = mensagem_final.replace(chave, str(valor))

    return mensagem_final


def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def normalizar_texto(texto):
    if texto is None:
        return ""

    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.replace(" ", "_")
    texto = texto.replace("/", "_")

    return texto


def get_valor(linha, *nomes):
    for nome in nomes:
        chave = normalizar_texto(nome)
        valor = linha.get(chave)

        if valor is not None and str(valor).strip() != "":
            return str(valor).strip()

    return ""


def converter_float(valor):
    if valor is None or str(valor).strip() == "":
        return 0.0

    valor = str(valor).strip()
    valor = valor.replace("R$", "").replace(" ", "")

    if "," in valor:
        valor = valor.replace(".", "").replace(",", ".")

    try:
        return float(valor)
    except ValueError:
        return 0.0


def converter_int(valor):
    if valor is None or str(valor).strip() == "":
        return 0

    try:
        return int(float(str(valor).replace(",", ".")))
    except ValueError:
        return 0


def ler_arquivo_planilha(arquivo):
    nome_arquivo = arquivo.filename.lower()

    if nome_arquivo.endswith(".csv"):
        conteudo = arquivo.stream.read().decode("utf-8-sig")

        try:
            dialect = csv.Sniffer().sniff(conteudo[:1024], delimiters=",;")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(io.StringIO(conteudo), dialect=dialect)

        linhas = []

        for row in reader:
            linha_normalizada = {}

            for chave, valor in row.items():
                linha_normalizada[normalizar_texto(chave)] = valor

            linhas.append(linha_normalizada)

        return linhas

    if nome_arquivo.endswith(".xlsx"):
        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active

        linhas = list(ws.iter_rows(values_only=True))

        if not linhas:
            return []

        cabecalhos = [normalizar_texto(celula) for celula in linhas[0]]
        resultado = []

        for linha in linhas[1:]:
            item = {}

            for index, valor in enumerate(linha):
                if index < len(cabecalhos):
                    item[cabecalhos[index]] = valor

            resultado.append(item)

        return resultado

    raise ValueError("Formato inválido. Envie um arquivo CSV ou XLSX.")

def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if "usuario_id" not in session:
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)
    return wrapper


def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if "usuario_id" not in session:
            return redirect(url_for("login"))

        if session.get("usuario_perfil") != "ADMIN":
            flash("Você não tem permissão para acessar esta área.")
            return redirect(url_for("vendas"))

        return func(*args, **kwargs)

    return wrapper


def usuario_logado():
    if "usuario_id" not in session:
        return None

    return {
        "id": session.get("usuario_id"),
        "nome": session.get("usuario_nome"),
        "perfil": session.get("usuario_perfil")
    }

def obter_configuracoes_loja():
    conn = get_db_connection()

    config = conn.execute("""
        SELECT
            id,
            nome_loja,
            telefone,
            endereco,
            cidade,
            instagram,
            mensagem_recibo,
            logo_path,
            login_logo_path
        FROM configuracoes_loja
        WHERE id = 1
    """).fetchone()

    conn.close()

    return config

def registrar_log(acao, entidade=None, entidade_id=None, descricao=None):
    try:
        conn = get_db_connection()

        conn.execute("""
            INSERT INTO logs_sistema (
                usuario_id,
                usuario_nome,
                acao,
                entidade,
                entidade_id,
                descricao
            ) VALUES (?, ?, ?, ?, ?, ?)
        """, (
            session.get("usuario_id"),
            session.get("usuario_nome"),
            acao,
            entidade,
            entidade_id,
            descricao
        ))

        conn.commit()
        conn.close()

    except Exception as erro:
        print(f"Erro ao registrar log: {erro}")

def usuario_atual():
    if "usuario_id" not in session:
        return None

    return {
        "id": session.get("usuario_id"),
        "nome": session.get("usuario_nome"),
        "perfil": session.get("usuario_perfil")
    }


def is_admin():
    usuario = usuario_atual()
    return usuario and usuario["perfil"] == "ADMIN"


def is_vendedor():
    usuario = usuario_atual()
    return usuario and usuario["perfil"] == "VENDEDOR"

def obter_modelo_whatsapp_por_codigo(conn, codigo):
    return conn.execute("""
        SELECT
            id,
            codigo,
            nome,
            mensagem
        FROM mensagens_whatsapp
        WHERE codigo = ?
          AND ativo = 1
        LIMIT 1
    """, (codigo,)).fetchone()


@app.route("/")
def index():
    if "usuario_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))

@app.route("/logs")
@admin_required
def logs():
    busca = request.args.get("busca", "").strip()
    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip()

    page = int(request.args.get("page", 1))
    per_page = 20

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    where_clauses = ["1 = 1"]
    params = []

    if busca:
        termo = f"%{busca}%"

        where_clauses.append("""
            (
                usuario_nome LIKE ?
                OR acao LIKE ?
                OR entidade LIKE ?
                OR descricao LIKE ?
            )
        """)

        params.extend([termo, termo, termo, termo])

    if data_inicio:
        where_clauses.append("date(created_at) >= date(?)")
        params.append(data_inicio)

    if data_fim:
        where_clauses.append("date(created_at) <= date(?)")
        params.append(data_fim)

    where_sql = " AND ".join(where_clauses)

    conn = get_db_connection()

    total_registros = conn.execute(f"""
        SELECT COUNT(*) AS total
        FROM logs_sistema
        WHERE {where_sql}
    """, params).fetchone()["total"]

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas == 0:
        total_paginas = 1

    if page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    logs_lista = conn.execute(f"""
        SELECT
            id,
            usuario_nome,
            acao,
            entidade,
            entidade_id,
            descricao,
            created_at
        FROM logs_sistema
        WHERE {where_sql}
        ORDER BY id DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()

    conn.close()

    return render_template(
        "logs.html",
        logs=logs_lista,
        busca=busca,
        data_inicio=data_inicio,
        data_fim=data_fim,
        page=page,
        total_paginas=total_paginas,
        total_registros=total_registros,
        usuario=usuario_logado()
    )

@app.route("/backup")
@admin_required
def backup():
    os.makedirs("backups", exist_ok=True)

    agora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nome_backup = f"crm_start_backup_{agora}.db"
    caminho_backup = os.path.join("backups", nome_backup)

    shutil.copy2(DATABASE, caminho_backup)

    registrar_log(
        "BACKUP_GERADO",
        "backup",
        None,
        f"Backup gerado: {nome_backup}."
    )

    flash("Backup gerado com sucesso.")
    return send_file(caminho_backup, as_attachment=True)

@app.route("/configuracoes", methods=["GET", "POST"])
@admin_required
def configuracoes():
    conn = get_db_connection()

    config = conn.execute("""
        SELECT
            id,
            nome_loja,
            telefone,
            endereco,
            cidade,
            instagram,
            mensagem_recibo,
            logo_path,
            login_logo_path
        FROM configuracoes_loja
        WHERE id = 1
    """).fetchone()

    if request.method == "POST":
        nome_loja = request.form.get("nome_loja")
        telefone = request.form.get("telefone")
        endereco = request.form.get("endereco")
        cidade = request.form.get("cidade")
        instagram = request.form.get("instagram")
        mensagem_recibo = request.form.get("mensagem_recibo")

        logo_path = config["logo_path"] if config and config["logo_path"] else None
        login_logo_path = config["login_logo_path"] if config and config["login_logo_path"] else None

        logo = request.files.get("logo")

        if logo and logo.filename:
            extensoes_permitidas = {"png", "jpg", "jpeg", "webp", "svg"}
            filename = secure_filename(logo.filename)
            extensao = filename.rsplit(".", 1)[-1].lower()

            if extensao not in extensoes_permitidas:
                flash("Formato de logo inválido. Use PNG, JPG, JPEG, WEBP ou SVG.")
                conn.close()
                return redirect(url_for("configuracoes"))

            os.makedirs("static/uploads", exist_ok=True)

            nome_arquivo = f"logo_start.{extensao}"
            caminho_arquivo = os.path.join("static", "uploads", nome_arquivo)

            logo.save(caminho_arquivo)

            logo_path = f"uploads/{nome_arquivo}"

        login_logo = request.files.get("login_logo")

        if login_logo and login_logo.filename:
            extensoes_permitidas = {"png", "jpg", "jpeg", "webp", "svg"}
            filename = secure_filename(login_logo.filename)
            extensao = filename.rsplit(".", 1)[-1].lower()

            if extensao not in extensoes_permitidas:
                flash("Formato da logo de login inválido. Use PNG, JPG, JPEG, WEBP ou SVG.")
                conn.close()
                return redirect(url_for("configuracoes"))

            os.makedirs("static/uploads", exist_ok=True)

            nome_arquivo = f"logo_login_start.{extensao}"
            caminho_arquivo = os.path.join("static", "uploads", nome_arquivo)

            login_logo.save(caminho_arquivo)

            login_logo_path = f"uploads/{nome_arquivo}"

        conn.execute("""
            INSERT INTO configuracoes_loja (
                id,
                nome_loja,
                telefone,
                endereco,
                cidade,
                instagram,
                mensagem_recibo,
                logo_path,
                login_logo_path,
                updated_at
            )
            VALUES (
                1, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP
            )
            ON CONFLICT(id) DO UPDATE SET
                nome_loja = excluded.nome_loja,
                telefone = excluded.telefone,
                endereco = excluded.endereco,
                cidade = excluded.cidade,
                instagram = excluded.instagram,
                mensagem_recibo = excluded.mensagem_recibo,
                logo_path = excluded.logo_path,
                login_logo_path = excluded.login_logo_path,
                updated_at = CURRENT_TIMESTAMP
        """, (
            nome_loja,
            telefone,
            endereco,
            cidade,
            instagram,
            mensagem_recibo,
            logo_path,
            login_logo_path
        ))

        conn.commit()
        conn.close()

        flash("Configurações atualizadas com sucesso.")
        return redirect(url_for("configuracoes"))

    conn.close()

    return render_template(
        "configuracoes.html",
        config=config,
        usuario=usuario_logado()
    )

@app.route("/usuarios", methods=["GET", "POST"])
@admin_required
def usuarios():
    conn = get_db_connection()

    if request.method == "POST":
        nome = request.form.get("nome")
        email = request.form.get("email")
        senha = request.form.get("senha")
        perfil = request.form.get("perfil")
        comissao_percentual = float(request.form.get("comissao_percentual") or 0)

        if perfil not in ["ADMIN", "VENDEDOR"]:
            flash("Perfil inválido.")
            conn.close()
            return redirect(url_for("usuarios"))

        senha_hash = generate_password_hash(senha)

        try:
            conn.execute("""
                INSERT INTO usuarios (
                    nome,
                    email,
                    senha_hash,
                    perfil,
                    comissao_percentual,
                    ativo
                ) VALUES (?, ?, ?, ?, ?, 1)
            """, (
                nome,
                email,
                senha_hash,
                perfil,
                comissao_percentual
            ))

            conn.commit()
            flash("Usuário cadastrado com sucesso.")

        except sqlite3.IntegrityError:
            flash("Erro: já existe um usuário com este e-mail.")

        conn.close()
        return redirect(url_for("usuarios"))

    usuarios_lista = conn.execute("""
        SELECT
            id,
            nome,
            email,
            perfil,
            comissao_percentual,
            ativo,
            created_at
        FROM usuarios
        ORDER BY nome ASC
    """).fetchall()

    conn.close()

    return render_template(
        "usuarios.html",
        usuarios=usuarios_lista,
        usuario=usuario_logado()
    )

@app.route("/usuarios/<int:usuario_id>/editar", methods=["GET", "POST"])
@admin_required
def usuario_editar(usuario_id):
    conn = get_db_connection()

    usuario_edicao = conn.execute("""
        SELECT
            id,
            nome,
            email,
            perfil,
            comissao_percentual,
            ativo
        FROM usuarios
        WHERE id = ?
    """, (usuario_id,)).fetchone()

    if not usuario_edicao:
        conn.close()
        flash("Usuário não encontrado.")
        return redirect(url_for("usuarios"))

    if request.method == "POST":
        nome = request.form.get("nome")
        email = request.form.get("email")
        perfil = request.form.get("perfil")
        comissao_percentual = float(request.form.get("comissao_percentual") or 0)
        nova_senha = request.form.get("nova_senha")
        ativo = int(request.form.get("ativo") or 0)

        if perfil not in ["ADMIN", "VENDEDOR"]:
            flash("Perfil inválido.")
            conn.close()
            return redirect(url_for("usuario_editar", usuario_id=usuario_id))

        try:
            if nova_senha:
                senha_hash = generate_password_hash(nova_senha)

                conn.execute("""
                    UPDATE usuarios
                    SET
                        nome = ?,
                        email = ?,
                        perfil = ?,
                        comissao_percentual = ?,
                        senha_hash = ?,
                        ativo = ?
                    WHERE id = ?
                """, (
                    nome,
                    email,
                    perfil,
                    comissao_percentual,
                    senha_hash,
                    ativo,
                    usuario_id
                ))
            else:
                conn.execute("""
                    UPDATE usuarios
                    SET
                        nome = ?,
                        email = ?,
                        perfil = ?,
                        comissao_percentual = ?,
                        ativo = ?
                    WHERE id = ?
                """, (
                    nome,
                    email,
                    perfil,
                    comissao_percentual,
                    ativo,
                    usuario_id
                ))

            conn.commit()
            conn.close()

            flash("Usuário atualizado com sucesso.")
            return redirect(url_for("usuarios"))

        except sqlite3.IntegrityError:
            conn.close()
            flash("Erro: já existe outro usuário com este e-mail.")
            return redirect(url_for("usuario_editar", usuario_id=usuario_id))

    conn.close()

    return render_template(
        "usuario_editar.html",
        usuario_edicao=usuario_edicao,
        usuario=usuario_logado()
    )

@app.route("/comissoes")
@admin_required
def comissoes():
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    hoje = date.today().isoformat()

    if not data_inicio:
        data_inicio = hoje

    if not data_fim:
        data_fim = hoje

    conn = get_db_connection()

    comissoes_lista = conn.execute("""
        SELECT
            u.id AS vendedor_id,
            u.nome AS vendedor_nome,
            u.email AS vendedor_email,
            u.comissao_percentual,
            COUNT(v.id) AS quantidade_vendas,
            COALESCE(SUM(v.valor_total), 0) AS total_vendido,
            COALESCE(SUM(v.lucro_total), 0) AS lucro_total,
            COALESCE(
                SUM(v.valor_total * (u.comissao_percentual / 100.0)),
                0
            ) AS valor_comissao
        FROM usuarios u
        LEFT JOIN vendas v 
            ON v.vendedor_id = u.id
            AND date(v.data_venda) BETWEEN date(?) AND date(?)
            AND v.status = 'CONCLUIDA'
        WHERE u.perfil = 'VENDEDOR'
          AND u.ativo = 1
        GROUP BY u.id
        ORDER BY total_vendido DESC, u.nome ASC
    """, (data_inicio, data_fim)).fetchall()

    resumo = conn.execute("""
        SELECT
            COALESCE(SUM(v.valor_total), 0) AS total_vendido,
            COALESCE(SUM(v.lucro_total), 0) AS lucro_total,
            COUNT(v.id) AS quantidade_vendas
        FROM vendas v
        WHERE date(v.data_venda) BETWEEN date(?) AND date(?)
          AND v.status = 'CONCLUIDA'
          AND v.vendedor_id IS NOT NULL
    """, (data_inicio, data_fim)).fetchone()

    total_comissoes = 0
    for item in comissoes_lista:
        total_comissoes += item["valor_comissao"]

    conn.close()

    return render_template(
        "comissoes.html",
        comissoes=comissoes_lista,
        resumo=resumo,
        total_comissoes=total_comissoes,
        data_inicio=data_inicio,
        data_fim=data_fim,
        usuario=usuario_logado()
    )

@app.route("/comissoes/exportar")
@admin_required
def exportar_comissoes():
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    hoje = date.today().isoformat()

    if not data_inicio:
        data_inicio = hoje

    if not data_fim:
        data_fim = hoje

    conn = get_db_connection()

    comissoes_lista = conn.execute("""
        SELECT
            u.nome AS vendedor_nome,
            u.email AS vendedor_email,
            u.comissao_percentual,
            COUNT(v.id) AS quantidade_vendas,
            COALESCE(SUM(v.valor_total), 0) AS total_vendido,
            COALESCE(SUM(v.lucro_total), 0) AS lucro_total,
            COALESCE(
                SUM(v.valor_total * (u.comissao_percentual / 100.0)),
                0
            ) AS valor_comissao
        FROM usuarios u
        LEFT JOIN vendas v 
            ON v.vendedor_id = u.id
            AND date(v.data_venda) BETWEEN date(?) AND date(?)
            AND v.status = 'CONCLUIDA'
        WHERE u.perfil = 'VENDEDOR'
          AND u.ativo = 1
        GROUP BY u.id
        ORDER BY total_vendido DESC, u.nome ASC
    """, (data_inicio, data_fim)).fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Comissões"

    ws.append([
        "Vendedor",
        "E-mail",
        "Quantidade de vendas",
        "Total vendido",
        "Lucro gerado",
        "Comissão %",
        "Valor comissão"
    ])

    total_geral_vendido = 0
    total_geral_lucro = 0
    total_geral_comissao = 0
    total_geral_vendas = 0

    for item in comissoes_lista:
        total_geral_vendido += item["total_vendido"]
        total_geral_lucro += item["lucro_total"]
        total_geral_comissao += item["valor_comissao"]
        total_geral_vendas += item["quantidade_vendas"]

        ws.append([
            item["vendedor_nome"],
            item["vendedor_email"],
            item["quantidade_vendas"],
            item["total_vendido"],
            item["lucro_total"],
            item["comissao_percentual"],
            item["valor_comissao"]
        ])

    ws.append([])
    ws.append([
        "TOTAL",
        "",
        total_geral_vendas,
        total_geral_vendido,
        total_geral_lucro,
        "",
        total_geral_comissao
    ])

    os.makedirs("exports", exist_ok=True)

    nome_arquivo = f"comissoes_{data_inicio}_a_{data_fim}.xlsx"
    caminho_arquivo = os.path.join("exports", nome_arquivo)

    wb.save(caminho_arquivo)

    return send_file(caminho_arquivo, as_attachment=True)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        senha = request.form.get("senha")

        conn = get_db_connection()
        usuario = conn.execute(
            "SELECT * FROM usuarios WHERE email = ? AND ativo = 1",
            (email,)
        ).fetchone()
        conn.close()

        if usuario and check_password_hash(usuario["senha_hash"], senha):
            session["usuario_id"] = usuario["id"]
            session["usuario_nome"] = usuario["nome"]
            session["usuario_perfil"] = usuario["perfil"]
            if usuario["perfil"] == "ADMIN":
                return redirect(url_for("dashboard"))

            return redirect(url_for("vendas"))

        flash("E-mail ou senha inválidos.")

    return render_template("login.html")


@app.route("/dashboard")
@admin_required
def dashboard():
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    per_page_dashboard = 10

    produtos_baixo_dashboard = buscar_produtos_dashboard(
        conn,
        "baixo",
        normalizar_pagina(request.args.get("page_produtos_baixo")),
        per_page_dashboard
    )

    produtos_sem_dashboard = buscar_produtos_dashboard(
        conn,
        "sem",
        normalizar_pagina(request.args.get("page_produtos_sem")),
        per_page_dashboard
    )

    vendas_hoje = conn.execute("""
        SELECT
            COUNT(*) AS quantidade,
            COALESCE(SUM(valor_total), 0) AS total,
            COALESCE(SUM(lucro_total), 0) AS lucro
        FROM vendas
        WHERE date(data_venda) = date('now', 'localtime')
          AND status = 'CONCLUIDA'
    """).fetchone()

    faturamento_mes = conn.execute("""
        SELECT
            COUNT(*) AS quantidade,
            COALESCE(SUM(valor_total), 0) AS total,
            COALESCE(SUM(lucro_total), 0) AS lucro
        FROM vendas
        WHERE strftime('%Y-%m', data_venda) = strftime('%Y-%m', 'now', 'localtime')
          AND status = 'CONCLUIDA'
    """).fetchone()

    ultimas_vendas = conn.execute("""
        SELECT
            v.id,
            v.data_venda,
            COALESCE(c.nome, 'Cliente não identificado') AS cliente_nome,
            v.vendedor,
            v.forma_pagamento,
            v.valor_total,
            v.lucro_total,
            v.status,
            (
                SELECT GROUP_CONCAT(pagamento.forma_pagamento, ' + ')
                FROM venda_pagamentos pagamento
                WHERE pagamento.venda_id = v.id
            ) AS pagamentos_descricao
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        ORDER BY v.id DESC
        LIMIT 5
    """).fetchall()

    produtos_alerta = conn.execute("""
        SELECT
            id,
            nome,
            categoria,
            marca,
            sku,
            estoque_atual,
            estoque_minimo
        FROM produtos
        WHERE ativo = 1
        AND estoque_atual <= estoque_minimo
        ORDER BY estoque_atual ASC, nome ASC
        LIMIT 5
    """).fetchall()

    clientes_reativar = buscar_clientes_dashboard(
        conn,
        "reativar",
        normalizar_pagina(request.args.get("page_reativar")),
        per_page_dashboard
    )

    clientes_vip_dashboard = buscar_clientes_dashboard(
        conn,
        "vip",
        normalizar_pagina(request.args.get("page_vip")),
        per_page_dashboard
    )

    clientes_sem_compra_dashboard = buscar_clientes_dashboard(
        conn,
        "sem_compra",
        normalizar_pagina(request.args.get("page_sem_compra")),
        per_page_dashboard
    )

    clientes_sem_tags_dashboard = buscar_clientes_dashboard(
        conn,
        "sem_tags",
        normalizar_pagina(request.args.get("page_sem_tags")),
        per_page_dashboard
    )

    clientes_recentes_30_dashboard = buscar_clientes_dashboard(
        conn,
        "recentes_30",
        normalizar_pagina(request.args.get("page_recentes_30")),
        per_page_dashboard
    )

    hoje_obj = date.today()
    ontem_obj = hoje_obj - timedelta(days=1)
    inicio_mes_obj = hoje_obj.replace(day=1)
    fim_mes_anterior_obj = inicio_mes_obj - timedelta(days=1)
    inicio_mes_anterior_obj = fim_mes_anterior_obj.replace(day=1)
    dias_comparacao_mes = min(hoje_obj.day, fim_mes_anterior_obj.day)
    fim_comparacao_mes_anterior_obj = (
        inicio_mes_anterior_obj + timedelta(days=dias_comparacao_mes - 1)
    )

    resumo_ontem = conn.execute("""
        SELECT
            COUNT(*) AS quantidade,
            COALESCE(SUM(valor_total), 0) AS total,
            COALESCE(SUM(lucro_total), 0) AS lucro
        FROM vendas
        WHERE date(data_venda) = date(?)
          AND status = 'CONCLUIDA'
    """, (ontem_obj.isoformat(),)).fetchone()

    resumo_mes_anterior = conn.execute("""
        SELECT
            COUNT(*) AS quantidade,
            COALESCE(SUM(valor_total), 0) AS total,
            COALESCE(SUM(lucro_total), 0) AS lucro
        FROM vendas
        WHERE date(data_venda) BETWEEN date(?) AND date(?)
          AND status = 'CONCLUIDA'
    """, (
        inicio_mes_anterior_obj.isoformat(),
        fim_comparacao_mes_anterior_obj.isoformat(),
    )).fetchone()

    inicio_serie_obj = hoje_obj - timedelta(days=13)
    serie_dashboard_rows = conn.execute("""
        SELECT
            date(data_venda) AS dia,
            COALESCE(SUM(valor_total), 0) AS total,
            COALESCE(SUM(lucro_total), 0) AS lucro
        FROM vendas
        WHERE date(data_venda) BETWEEN date(?) AND date(?)
          AND status = 'CONCLUIDA'
        GROUP BY date(data_venda)
        ORDER BY date(data_venda)
    """, (
        inicio_serie_obj.isoformat(),
        hoje_obj.isoformat(),
    )).fetchall()

    pagamentos_hoje_rows = conn.execute("""
        SELECT
            vp.forma_pagamento,
            COUNT(DISTINCT vp.venda_id) AS quantidade,
            COALESCE(SUM(vp.valor), 0) AS total
        FROM venda_pagamentos vp
        INNER JOIN vendas v ON v.id = vp.venda_id
        WHERE date(v.data_venda) = date('now', 'localtime')
          AND v.status = 'CONCLUIDA'
        GROUP BY vp.forma_pagamento
        ORDER BY total DESC
    """).fetchall()

    caixa_aberto_dashboard = obter_caixa_aberto(conn)
    resumo_caixa_dashboard = (
        calcular_resumo_caixa(conn, caixa_aberto_dashboard["id"])
        if caixa_aberto_dashboard
        else None
    )

    def calcular_variacao_dashboard(valor_atual, valor_anterior):
        valor_atual = float(valor_atual or 0)
        valor_anterior = float(valor_anterior or 0)

        if abs(valor_anterior) <= 0.009:
            return None if abs(valor_atual) <= 0.009 else 100.0

        return round(((valor_atual - valor_anterior) / abs(valor_anterior)) * 100, 1)

    ticket_medio_hoje = (
        float(vendas_hoje["total"] or 0) / int(vendas_hoje["quantidade"] or 0)
        if int(vendas_hoje["quantidade"] or 0) > 0
        else 0
    )
    margem_mes = (
        float(faturamento_mes["lucro"] or 0)
        / float(faturamento_mes["total"] or 0)
        * 100
        if float(faturamento_mes["total"] or 0) > 0
        else 0
    )
    variacoes_dashboard = {
        "vendas_hoje": calcular_variacao_dashboard(
            vendas_hoje["quantidade"],
            resumo_ontem["quantidade"],
        ),
        "faturamento_hoje": calcular_variacao_dashboard(
            vendas_hoje["total"],
            resumo_ontem["total"],
        ),
        "faturamento_mes": calcular_variacao_dashboard(
            faturamento_mes["total"],
            resumo_mes_anterior["total"],
        ),
        "lucro_mes": calcular_variacao_dashboard(
            faturamento_mes["lucro"],
            resumo_mes_anterior["lucro"],
        ),
    }

    serie_por_dia = {
        row["dia"]: {
            "total": float(row["total"] or 0),
            "lucro": float(row["lucro"] or 0),
        }
        for row in serie_dashboard_rows
    }
    serie_dashboard = []

    for deslocamento in range(14):
        dia_obj = inicio_serie_obj + timedelta(days=deslocamento)
        valores = serie_por_dia.get(dia_obj.isoformat(), {})
        serie_dashboard.append({
            "label": dia_obj.strftime("%d/%m"),
            "total": valores.get("total", 0),
            "lucro": valores.get("lucro", 0),
        })

    maior_valor_dashboard = max(
        [
            max(item["total"], item["lucro"], 0)
            for item in serie_dashboard
        ]
        or [0]
    )

    for item in serie_dashboard:
        if maior_valor_dashboard > 0:
            item["altura_total"] = (
                max(4, round(item["total"] / maior_valor_dashboard * 100, 2))
                if item["total"] > 0
                else 0
            )
            item["altura_lucro"] = (
                max(4, round(max(item["lucro"], 0) / maior_valor_dashboard * 100, 2))
                if item["lucro"] > 0
                else 0
            )
        else:
            item["altura_total"] = 0
            item["altura_lucro"] = 0

    total_pagamentos_hoje = sum(
        float(item["total"] or 0)
        for item in pagamentos_hoje_rows
    )
    pagamentos_hoje = [
        {
            **dict(item),
            "percentual": (
                round(float(item["total"] or 0) / total_pagamentos_hoje * 100, 1)
                if total_pagamentos_hoje > 0
                else 0
            ),
        }
        for item in pagamentos_hoje_rows
    ]

    modal_aberto = request.args.get("modal", "")

    conn.close()

    return render_template(
        "dashboard.html",
        usuario_nome=session.get("usuario_nome"),
        vendas_hoje=vendas_hoje["total"],
        quantidade_vendas_hoje=vendas_hoje["quantidade"],
        ticket_medio_hoje=ticket_medio_hoje,
        faturamento_mes=faturamento_mes["total"],
        lucro_mes=faturamento_mes["lucro"],
        margem_mes=margem_mes,
        variacoes_dashboard=variacoes_dashboard,
        serie_dashboard=serie_dashboard,
        pagamentos_hoje=pagamentos_hoje,
        caixa_aberto_dashboard=caixa_aberto_dashboard,
        resumo_caixa_dashboard=resumo_caixa_dashboard,
        ultimas_vendas=ultimas_vendas,
        produtos_alerta=produtos_alerta,
        clientes_reativar=clientes_reativar,
        clientes_vip_dashboard=clientes_vip_dashboard,
        clientes_sem_compra_dashboard=clientes_sem_compra_dashboard,
        clientes_sem_tags_dashboard=clientes_sem_tags_dashboard,
        clientes_recentes_30_dashboard=clientes_recentes_30_dashboard,
        modal_aberto=modal_aberto,
        produtos_baixo_dashboard=produtos_baixo_dashboard,
        produtos_sem_dashboard=produtos_sem_dashboard
    )

@app.route("/clientes", methods=["GET", "POST"])
@login_required
def clientes():
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    if request.method == "POST":
        nome = request.form.get("nome")
        telefone = request.form.get("telefone")
        endereco_completo = request.form.get("endereco_completo")
        observacoes = request.form.get("observacoes")
        tags_texto = request.form.get("tags")

        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO clientes (
                nome,
                telefone,
                endereco_completo,
                observacoes
            ) VALUES (?, ?, ?, ?)
        """, (
            nome,
            telefone,
            endereco_completo,
            observacoes
        ))

        cliente_id = cursor.lastrowid

        if tags_texto:
            tags = [tag.strip() for tag in tags_texto.split(",") if tag.strip()]

            for tag in tags:
                cursor.execute("""
                    INSERT INTO cliente_tags (
                        cliente_id,
                        tag
                    ) VALUES (?, ?)
                """, (
                    cliente_id,
                    tag
                ))

        conn.commit()
        flash("Cliente cadastrado com sucesso.")
        return redirect(url_for("clientes"))

    busca = request.args.get("busca", "").strip()
    status = request.args.get("status", "ativos")
    page = int(request.args.get("page", 1))
    per_page = 10

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    where_clauses = []
    params = []

    if status == "inativos":
        where_clauses.append("c.ativo = 0")
    elif status == "todos":
        where_clauses.append("c.ativo IN (0, 1)")
    else:
        where_clauses.append("c.ativo = 1")

    if busca:
        termo = f"%{busca}%"
        where_clauses.append("""
            (
                c.nome LIKE ?
                OR c.telefone LIKE ?
                OR c.endereco_completo LIKE ?
                OR ct.tag LIKE ?
            )
        """)
        params.extend([termo, termo, termo, termo])

    where_sql = " AND ".join(where_clauses)

    resumo_clientes = conn.execute("""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN ativo = 1 THEN 1 ELSE 0 END) AS ativos,
            SUM(CASE WHEN ativo = 0 THEN 1 ELSE 0 END) AS inativos,
            SUM(
                CASE
                    WHEN ativo = 1
                     AND date(created_at) >= date('now', 'localtime', '-30 days')
                    THEN 1
                    ELSE 0
                END
            ) AS novos_30_dias
        FROM clientes
    """).fetchone()

    total_registros = conn.execute(f"""
        SELECT COUNT(DISTINCT c.id) AS total
        FROM clientes c
        LEFT JOIN cliente_tags ct ON ct.cliente_id = c.id
        WHERE {where_sql}
    """, params).fetchone()["total"]

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas == 0:
        total_paginas = 1

    if page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    clientes_lista = conn.execute(f"""
        SELECT 
            c.id,
            c.nome,
            c.telefone,
            c.endereco_completo,
            c.observacoes,
            c.ativo,
            c.created_at,
            GROUP_CONCAT(DISTINCT ct.tag) AS tags,
            COALESCE(rv.quantidade_compras, 0) AS quantidade_compras,
            COALESCE(rv.total_comprado, 0) AS total_comprado,
            rv.ultima_compra
        FROM clientes c
        LEFT JOIN cliente_tags ct ON ct.cliente_id = c.id
        LEFT JOIN (
            SELECT
                cliente_id,
                COUNT(*) AS quantidade_compras,
                COALESCE(SUM(valor_total), 0) AS total_comprado,
                MAX(data_venda) AS ultima_compra
            FROM vendas
            WHERE status = 'CONCLUIDA'
              AND cliente_id IS NOT NULL
            GROUP BY cliente_id
        ) rv ON rv.cliente_id = c.id
        WHERE {where_sql}
        GROUP BY c.id
        ORDER BY c.nome ASC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()

    conn.close()

    return render_template(
        "clientes.html",
        clientes=clientes_lista,
        busca=busca,
        status=status,
        page=page,
        total_paginas=total_paginas,
        total_registros=total_registros,
        resumo_clientes=resumo_clientes,
        usuario=usuario_logado()
    )

@app.route("/clientes/importar", methods=["POST"])
@admin_required
def importar_clientes():
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    arquivo = request.files.get("arquivo")

    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo CSV ou Excel.")
        return redirect(url_for("clientes"))

    try:
        linhas = ler_arquivo_planilha(arquivo)
    except Exception as erro:
        flash(f"Erro ao ler arquivo: {erro}")
        return redirect(url_for("clientes"))

    conn = get_db_connection()
    cursor = conn.cursor()

    importados = 0
    atualizados = 0
    ignorados = 0

    for linha in linhas:
        nome = get_valor(linha, "nome", "cliente")
        telefone = get_valor(linha, "telefone", "whatsapp", "celular")
        endereco_completo = get_valor(linha, "endereco_completo", "endereco", "endereço")
        observacoes = get_valor(linha, "observacoes", "observações", "obs")
        tags_texto = get_valor(linha, "tags", "interesses", "preferencias", "preferências")

        if not nome or not telefone:
            ignorados += 1
            continue

        cliente_existente = cursor.execute("""
            SELECT id
            FROM clientes
            WHERE telefone = ?
            LIMIT 1
        """, (telefone,)).fetchone()

        if cliente_existente:
            cliente_id = cliente_existente["id"]

            cursor.execute("""
                UPDATE clientes
                SET
                    nome = ?,
                    endereco_completo = ?,
                    observacoes = ?,
                    ativo = 1,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                nome,
                endereco_completo,
                observacoes,
                cliente_id
            ))

            atualizados += 1

        else:
            cursor.execute("""
                INSERT INTO clientes (
                    nome,
                    telefone,
                    endereco_completo,
                    observacoes,
                    ativo
                ) VALUES (?, ?, ?, ?, 1)
            """, (
                nome,
                telefone,
                endereco_completo,
                observacoes
            ))

            cliente_id = cursor.lastrowid
            importados += 1

        if tags_texto:
            cursor.execute("""
                DELETE FROM cliente_tags
                WHERE cliente_id = ?
            """, (cliente_id,))

            tags = [tag.strip() for tag in tags_texto.split(",") if tag.strip()]

            for tag in tags:
                cursor.execute("""
                    INSERT INTO cliente_tags (
                        cliente_id,
                        tag
                    ) VALUES (?, ?)
                """, (
                    cliente_id,
                    tag
                ))

    conn.commit()
    conn.close()

    flash(f"Importação concluída. Novos: {importados}. Atualizados: {atualizados}. Ignorados: {ignorados}.")
    return redirect(url_for("clientes"))

@app.route("/clientes/modelo")
@admin_required
def baixar_modelo_clientes():
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Clientes"

    ws.append([
        "nome",
        "telefone",
        "endereco_completo",
        "tags",
        "observacoes"
    ])

    ws.append([
        "Maria Silva",
        "88999999999",
        "Rua Teste, 100",
        "iPhone, Capinhas, Películas",
        "Cliente recorrente"
    ])

    ws.append([
        "João Santos",
        "88888888888",
        "Rua Central, 50",
        "Samsung, Carregadores",
        ""
    ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = max_length + 4

    os.makedirs("exports", exist_ok=True)

    caminho_arquivo = os.path.join("exports", "modelo_importacao_clientes.xlsx")
    wb.save(caminho_arquivo)

    return send_file(caminho_arquivo, as_attachment=True)

@app.route("/clientes/<int:cliente_id>")
@login_required
def cliente_detalhe(cliente_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    cliente = conn.execute("""
        SELECT
            id,
            nome,
            telefone,
            endereco_completo,
            observacoes,
            created_at
        FROM clientes
        WHERE id = ? AND ativo = 1
    """, (cliente_id,)).fetchone()

    if not cliente:
        conn.close()
        flash("Cliente não encontrado.")
        return redirect(url_for("clientes"))

    tags = conn.execute("""
        SELECT tag
        FROM cliente_tags
        WHERE cliente_id = ?
        ORDER BY tag ASC
    """, (cliente_id,)).fetchall()

    resumo = conn.execute("""
        SELECT
            COUNT(*) AS quantidade_compras,
            COALESCE(SUM(valor_total), 0) AS total_comprado,
            COALESCE(SUM(lucro_total), 0) AS lucro_gerado,
            MAX(data_venda) AS ultima_compra
        FROM vendas
        WHERE cliente_id = ?
          AND status = 'CONCLUIDA'
    """, (cliente_id,)).fetchone()

    quantidade_compras = int(resumo["quantidade_compras"] or 0)
    total_comprado = float(resumo["total_comprado"] or 0)
    lucro_gerado = float(resumo["lucro_gerado"] or 0)
    ticket_medio = (
        total_comprado / quantidade_compras
        if quantidade_compras > 0
        else 0
    )
    margem_cliente = (
        lucro_gerado / total_comprado * 100
        if total_comprado > 0
        else 0
    )

    vendas = conn.execute("""
        SELECT
            id,
            data_venda,
            vendedor,
            forma_pagamento,
            desconto_total,
            valor_total,
            custo_total,
            lucro_total,
            observacoes,
            status,
            (
                SELECT GROUP_CONCAT(vp.forma_pagamento, ' + ')
                FROM venda_pagamentos vp
                WHERE vp.venda_id = vendas.id
            ) AS pagamentos_descricao
        FROM vendas
        WHERE cliente_id = ?
        ORDER BY data_venda DESC
    """, (cliente_id,)).fetchall()

    produtos_comprados = conn.execute("""
        SELECT
            p.nome AS produto_nome,
            p.sku,
            p.categoria,
            SUM(vi.quantidade) AS quantidade_total,
            SUM(vi.subtotal) AS total_gasto
        FROM venda_itens vi
        INNER JOIN vendas v ON v.id = vi.venda_id
        INNER JOIN produtos p ON p.id = vi.produto_id
        WHERE v.cliente_id = ?
          AND v.status = 'CONCLUIDA'
        GROUP BY p.id
        ORDER BY quantidade_total DESC, total_gasto DESC
    """, (cliente_id,)).fetchall()

    categorias_preferidas = conn.execute("""
        SELECT
            p.categoria,
            SUM(vi.quantidade) AS quantidade_total,
            SUM(vi.subtotal) AS total_gasto
        FROM venda_itens vi
        INNER JOIN vendas v ON v.id = vi.venda_id
        INNER JOIN produtos p ON p.id = vi.produto_id
        WHERE v.cliente_id = ?
          AND v.status = 'CONCLUIDA'
        GROUP BY p.categoria
        ORDER BY quantidade_total DESC, total_gasto DESC
    """, (cliente_id,)).fetchall()

    total_categorias = sum(
        float(categoria["total_gasto"] or 0)
        for categoria in categorias_preferidas
    )

    modelos_whatsapp = conn.execute("""
        SELECT
            id,
            nome,
            categoria
        FROM mensagens_whatsapp
        WHERE ativo = 1
        ORDER BY nome ASC
    """).fetchall()

    conn.close()

    return render_template(
        "cliente_detalhe.html",
        cliente=cliente,
        tags=tags,
        resumo=resumo,
        ticket_medio=ticket_medio,
        margem_cliente=margem_cliente,
        vendas=vendas,
        produtos_comprados=produtos_comprados,
        categorias_preferidas=categorias_preferidas,
        total_categorias=total_categorias,
        modelos_whatsapp=modelos_whatsapp,
        usuario=usuario_logado()
    )

@app.route("/clientes/<int:cliente_id>/editar", methods=["GET", "POST"])
@login_required
def cliente_editar(cliente_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    cliente = conn.execute("""
        SELECT
            id,
            nome,
            telefone,
            endereco_completo,
            observacoes,
            created_at
        FROM clientes
        WHERE id = ? AND ativo = 1
    """, (cliente_id,)).fetchone()

    if not cliente:
        conn.close()
        flash("Cliente não encontrado.")
        return redirect(url_for("clientes"))

    if request.method == "POST":
        nome = request.form.get("nome")
        telefone = request.form.get("telefone")
        endereco_completo = request.form.get("endereco_completo")
        observacoes = request.form.get("observacoes")
        tags_texto = request.form.get("tags")

        cursor = conn.cursor()

        cursor.execute("""
            UPDATE clientes
            SET
                nome = ?,
                telefone = ?,
                endereco_completo = ?,
                observacoes = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (
            nome,
            telefone,
            endereco_completo,
            observacoes,
            cliente_id
        ))

        cursor.execute("""
            DELETE FROM cliente_tags
            WHERE cliente_id = ?
        """, (cliente_id,))

        if tags_texto:
            tags = [tag.strip() for tag in tags_texto.split(",") if tag.strip()]

            for tag in tags:
                cursor.execute("""
                    INSERT INTO cliente_tags (
                        cliente_id,
                        tag
                    ) VALUES (?, ?)
                """, (
                    cliente_id,
                    tag
                ))

        conn.commit()
        conn.close()

        flash("Cliente atualizado com sucesso.")
        return redirect(url_for("cliente_detalhe", cliente_id=cliente_id))

    tags = conn.execute("""
        SELECT tag
        FROM cliente_tags
        WHERE cliente_id = ?
        ORDER BY tag ASC
    """, (cliente_id,)).fetchall()

    tags_texto = ", ".join([tag["tag"] for tag in tags])
    quantidade_tags = len(tags)

    conn.close()

    return render_template(
        "cliente_editar.html",
        cliente=cliente,
        tags_texto=tags_texto,
        quantidade_tags=quantidade_tags,
        usuario=usuario_logado()
    )

@app.route("/clientes/<int:cliente_id>/inativar", methods=["POST"])
@admin_required
def cliente_inativar(cliente_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    cliente = conn.execute("""
        SELECT id
        FROM clientes
        WHERE id = ? AND ativo = 1
    """, (cliente_id,)).fetchone()

    if not cliente:
        conn.close()
        flash("Cliente não encontrado.")
        return redirect(url_for("clientes"))

    conn.execute("""
        UPDATE clientes
        SET ativo = 0,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (cliente_id,))

    conn.commit()
    conn.close()

    registrar_log(
        "CLIENTE_INATIVADO",
        "clientes",
        cliente_id,
        f"Cliente #{cliente_id} foi inativado."
    )

    flash("Cliente inativado com sucesso.")
    return redirect(url_for("clientes"))

@app.route("/clientes/<int:cliente_id>/reativar", methods=["POST"])
@admin_required
def cliente_reativar(cliente_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    cliente = conn.execute("""
        SELECT id
        FROM clientes
        WHERE id = ? AND ativo = 0
    """, (cliente_id,)).fetchone()

    if not cliente:
        conn.close()
        flash("Cliente inativo não encontrado.")
        return redirect(url_for("clientes", status="inativos"))

    conn.execute("""
        UPDATE clientes
        SET ativo = 1,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (cliente_id,))

    conn.commit()
    conn.close()

    registrar_log(
        "CLIENTE_REATIVADO",
        "clientes",
        cliente_id,
        f"Cliente #{cliente_id} foi reativado."
    )

    flash("Cliente reativado com sucesso.")
    return redirect(url_for("clientes", status="ativos"))

@app.route("/mensagens-whatsapp", methods=["GET", "POST"])
@admin_required
def mensagens_whatsapp():
    conn = get_db_connection()

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        categoria = request.form.get("categoria", "").strip()
        mensagem = request.form.get("mensagem", "").strip()

        if not nome or not mensagem:
            flash("Informe o nome e a mensagem.")
            conn.close()
            return redirect(url_for("mensagens_whatsapp"))

        conn.execute("""
            INSERT INTO mensagens_whatsapp (
                nome,
                categoria,
                mensagem,
                ativo
            ) VALUES (?, ?, ?, 1)
        """, (
            nome,
            categoria,
            mensagem
        ))

        conn.commit()
        conn.close()

        registrar_log(
            "MODELO_WHATSAPP_CRIADO",
            "mensagens_whatsapp",
            None,
            f"Modelo de WhatsApp criado: {nome}."
        )

        flash("Modelo de mensagem cadastrado com sucesso.")
        return redirect(url_for("mensagens_whatsapp"))

    mensagens = conn.execute("""
        SELECT
            id,
            codigo, 
            nome,
            categoria,
            mensagem,
            ativo,
            created_at
        FROM mensagens_whatsapp
        ORDER BY id DESC
    """).fetchall()

    conn.close()

    return render_template(
        "mensagens_whatsapp.html",
        mensagens=mensagens,
        usuario=usuario_logado()
    )

@app.route("/mensagens-whatsapp/<int:mensagem_id>/editar", methods=["GET", "POST"])
@admin_required
def mensagem_whatsapp_editar(mensagem_id):
    conn = get_db_connection()

    modelo = conn.execute("""
        SELECT
            id,
            codigo,
            nome,
            categoria,
            mensagem,
            ativo
        FROM mensagens_whatsapp
        WHERE id = ?
    """, (mensagem_id,)).fetchone()

    if not modelo:
        conn.close()
        flash("Modelo de mensagem não encontrado.")
        return redirect(url_for("mensagens_whatsapp"))

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip().upper()
        nome = request.form.get("nome", "").strip()
        categoria = request.form.get("categoria", "").strip()
        mensagem = request.form.get("mensagem", "").strip()
        ativo = int(request.form.get("ativo") or 0)

        if not nome or not mensagem:
            conn.close()
            flash("Informe o nome e a mensagem.")
            return redirect(url_for("mensagem_whatsapp_editar", mensagem_id=mensagem_id))

        conn.execute("""
            UPDATE mensagens_whatsapp
            SET
                codigo = ?,
                nome = ?,
                categoria = ?,
                mensagem = ?,
                ativo = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (
            codigo,
            nome,
            categoria,
            mensagem,
            ativo,
            mensagem_id
        ))

        conn.commit()
        conn.close()

        registrar_log(
            "MODELO_WHATSAPP_ATUALIZADO",
            "mensagens_whatsapp",
            mensagem_id,
            f"Modelo de WhatsApp atualizado: {nome}."
        )

        flash("Modelo de mensagem atualizado com sucesso.")
        return redirect(url_for("mensagens_whatsapp"))

    conn.close()

    return render_template(
        "mensagem_whatsapp_editar.html",
        modelo=modelo,
        usuario=usuario_logado()
    )

@app.route("/mensagens-whatsapp/<int:mensagem_id>/alternar-status", methods=["POST"])
@admin_required
def mensagem_whatsapp_alternar_status(mensagem_id):
    conn = get_db_connection()

    modelo = conn.execute("""
        SELECT
            id,
            nome,
            ativo
        FROM mensagens_whatsapp
        WHERE id = ?
    """, (mensagem_id,)).fetchone()

    if not modelo:
        conn.close()
        flash("Modelo de mensagem não encontrado.")
        return redirect(url_for("mensagens_whatsapp"))

    novo_status = 0 if modelo["ativo"] == 1 else 1

    conn.execute("""
        UPDATE mensagens_whatsapp
        SET
            ativo = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (
        novo_status,
        mensagem_id
    ))

    conn.commit()
    conn.close()

    acao = "reativado" if novo_status == 1 else "inativado"

    registrar_log(
        "MODELO_WHATSAPP_STATUS_ALTERADO",
        "mensagens_whatsapp",
        mensagem_id,
        f"Modelo de WhatsApp {acao}: {modelo['nome']}."
    )

    flash(f"Modelo de mensagem {acao} com sucesso.")
    return redirect(url_for("mensagens_whatsapp"))

@app.route("/clientes/<int:cliente_id>/whatsapp", methods=["POST"])
@login_required
def cliente_whatsapp(cliente_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    modelo_id = request.form.get("modelo_id")

    if not modelo_id:
        flash("Selecione um modelo de mensagem.")
        return redirect(url_for("cliente_detalhe", cliente_id=cliente_id))

    conn = get_db_connection()

    cliente = conn.execute("""
        SELECT
            id,
            nome,
            telefone
        FROM clientes
        WHERE id = ? AND ativo = 1
    """, (cliente_id,)).fetchone()

    if not cliente:
        conn.close()
        flash("Cliente não encontrado.")
        return redirect(url_for("clientes"))

    modelo = conn.execute("""
        SELECT
            id,
            nome,
            mensagem
        FROM mensagens_whatsapp
        WHERE id = ?
          AND ativo = 1
    """, (modelo_id,)).fetchone()

    if not modelo:
        conn.close()
        flash("Modelo de mensagem não encontrado.")
        return redirect(url_for("cliente_detalhe", cliente_id=cliente_id))

    config = conn.execute("""
        SELECT
            nome_loja,
            telefone,
            instagram
        FROM configuracoes_loja
        WHERE id = 1
    """).fetchone()

    resumo = conn.execute("""
        SELECT
            MAX(data_venda) AS ultima_compra,
            COALESCE(SUM(valor_total), 0) AS total_comprado
        FROM vendas
        WHERE cliente_id = ?
          AND status = 'CONCLUIDA'
    """, (cliente_id,)).fetchone()

    tags_rows = conn.execute("""
        SELECT tag
        FROM cliente_tags
        WHERE cliente_id = ?
        ORDER BY tag ASC
    """, (cliente_id,)).fetchall()

    conn.close()

    tags = [item["tag"] for item in tags_rows]

    telefone = "".join([c for c in cliente["telefone"] if c.isdigit()])

    if telefone.startswith("0"):
        telefone = telefone[1:]

    if not telefone.startswith("55"):
        telefone = "55" + telefone

    mensagem_final = aplicar_variaveis_mensagem(
        modelo["mensagem"],
        cliente,
        config,
        resumo,
        tags
    )

    from urllib.parse import quote

    mensagem_url = quote(mensagem_final)

    registrar_log(
        "WHATSAPP_GERADO",
        "clientes",
        cliente_id,
        f"Mensagem WhatsApp gerada para {cliente['nome']} usando o modelo {modelo['nome']}."
    )

    return redirect(f"https://wa.me/{telefone}?text={mensagem_url}")

@app.route("/produtos", methods=["GET", "POST"])
@admin_required
def produtos():
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    if request.method == "POST":
        nome = request.form.get("nome")
        categoria = request.form.get("categoria")
        marca = request.form.get("marca")
        sku = request.form.get("sku")
        preco_custo = request.form.get("preco_custo") or 0
        preco_venda = request.form.get("preco_venda") or 0
        estoque_atual = request.form.get("estoque_atual") or 0
        estoque_minimo = request.form.get("estoque_minimo") or 0

        try:
            cursor = conn.cursor()

            cursor.execute("""
                INSERT INTO produtos (
                    nome,
                    categoria,
                    marca,
                    sku,
                    preco_custo,
                    preco_venda,
                    estoque_atual,
                    estoque_minimo
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                nome,
                categoria,
                marca,
                sku,
                float(preco_custo),
                float(preco_venda),
                int(estoque_atual),
                int(estoque_minimo)
            ))

            produto_id = cursor.lastrowid

            cursor.execute("""
                INSERT INTO movimentacoes_estoque (
                    produto_id,
                    tipo,
                    quantidade,
                    estoque_anterior,
                    estoque_atual,
                    observacoes
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                produto_id,
                "ENTRADA_INICIAL",
                int(estoque_atual),
                0,
                int(estoque_atual),
                "Estoque inicial do produto"
            ))

            conn.commit()
            flash("Produto cadastrado com sucesso.")

        except sqlite3.IntegrityError:
            flash("Erro: já existe um produto com este SKU/código.")

        return redirect(url_for("produtos"))

    busca = request.args.get("busca", "").strip()
    status = request.args.get("status", "ativos")
    page = int(request.args.get("page", 1))
    per_page = 10

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    where_clauses = []
    params = []

    if status == "inativos":
        where_clauses.append("ativo = 0")
    elif status == "todos":
        where_clauses.append("ativo IN (0, 1)")
    else:
        where_clauses.append("ativo = 1")

    if busca:
        termo = f"%{busca}%"
        where_clauses.append("""
            (
                nome LIKE ?
                OR categoria LIKE ?
                OR marca LIKE ?
                OR sku LIKE ?
            )
        """)
        params.extend([termo, termo, termo, termo])

    where_sql = " AND ".join(where_clauses)

    resumo_produtos = conn.execute("""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN ativo = 1 THEN 1 ELSE 0 END) AS ativos,
            SUM(
                CASE
                    WHEN ativo = 1
                    THEN estoque_atual
                    ELSE 0
                END
            ) AS unidades_estoque,
            SUM(
                CASE
                    WHEN ativo = 1
                     AND estoque_atual > 0
                     AND estoque_atual <= estoque_minimo
                    THEN 1
                    ELSE 0
                END
            ) AS estoque_baixo,
            SUM(
                CASE
                    WHEN ativo = 1 AND estoque_atual = 0
                    THEN 1
                    ELSE 0
                END
            ) AS sem_estoque,
            SUM(
                CASE
                    WHEN ativo = 1
                    THEN estoque_atual * preco_custo
                    ELSE 0
                END
            ) AS valor_estoque
        FROM produtos
    """).fetchone()

    total_registros = conn.execute(f"""
        SELECT COUNT(*) AS total
        FROM produtos
        WHERE {where_sql}
    """, params).fetchone()["total"]

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas == 0:
        total_paginas = 1

    if page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    produtos_lista = conn.execute(f"""
        SELECT 
            id,
            nome,
            categoria,
            marca,
            sku,
            preco_custo,
            preco_venda,
            estoque_atual,
            estoque_minimo,
            ativo,
            created_at
        FROM produtos
        WHERE {where_sql}
        ORDER BY nome ASC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()

    conn.close()

    return render_template(
        "produtos.html",
        produtos=produtos_lista,
        busca=busca,
        status=status,
        page=page,
        total_paginas=total_paginas,
        total_registros=total_registros,
        resumo_produtos=resumo_produtos,
        usuario=usuario_logado()
    )

def gerar_planilha_produtos_ignorados(registros_ignorados, importados, atualizados):
    os.makedirs("exports", exist_ok=True)

    agora = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"produtos_ignorados_{agora}.xlsx"
    caminho_arquivo = os.path.join("exports", nome_arquivo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos Ignorados"

    ws.append([
        "Resumo",
        f"Novos: {importados}",
        f"Atualizados: {atualizados}",
        f"Ignorados: {len(registros_ignorados)}"
    ])
    ws.append([])

    ws.append([
        "Linha",
        "Nome",
        "SKU/Código",
        "Categoria",
        "Marca",
        "Preço custo",
        "Preço venda",
        "Tipo de atualização",
        "Estoque atual",
        "Estoque mínimo",
        "Motivo"
    ])

    for item in registros_ignorados:
        ws.append([
            item.get("linha"),
            item.get("nome"),
            item.get("sku"),
            item.get("categoria"),
            item.get("marca"),
            item.get("preco_custo"),
            item.get("preco_venda"),
            item.get("tipo_atualizacao"),
            item.get("estoque_atual"),
            item.get("estoque_minimo"),
            item.get("motivo")
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 4, 60)

    wb.save(caminho_arquivo)

    return caminho_arquivo, nome_arquivo


def gerar_historico_importacao_produtos(nome_arquivo_origem, registros):
    os.makedirs(HISTORICO_IMPORTACOES_PRODUTOS_DIR, exist_ok=True)

    data_importacao = datetime.now(timezone.utc)
    nome_historico = (
        f"importacao_produtos_{data_importacao.strftime('%Y%m%d_%H%M%S_%f')}.csv"
    )
    caminho_historico = os.path.join(
        HISTORICO_IMPORTACOES_PRODUTOS_DIR,
        nome_historico
    )

    campos = [
        "data_importacao",
        "usuario",
        "arquivo_origem",
        "linha",
        "nome",
        "sku",
        "tipo_atualizacao",
        "quantidade_informada",
        "estoque_anterior",
        "estoque_final",
        "resultado",
        "motivo"
    ]

    if not registros:
        registros = [{
            "linha": "",
            "nome": "",
            "sku": "",
            "tipo_atualizacao": "",
            "quantidade_informada": "",
            "estoque_anterior": "",
            "estoque_final": "",
            "resultado": "SEM_REGISTROS",
            "motivo": "O arquivo não continha linhas de produtos."
        }]

    with open(caminho_historico, "w", newline="", encoding="utf-8-sig") as arquivo_csv:
        writer = csv.DictWriter(arquivo_csv, fieldnames=campos, delimiter=";")
        writer.writeheader()

        for registro in registros:
            writer.writerow({
                "data_importacao": data_importacao.strftime("%Y-%m-%d %H:%M:%S"),
                "usuario": session.get("usuario_nome") or "-",
                "arquivo_origem": nome_arquivo_origem,
                "linha": registro.get("linha", ""),
                "nome": registro.get("nome", ""),
                "sku": registro.get("sku", ""),
                "tipo_atualizacao": registro.get("tipo_atualizacao", ""),
                "quantidade_informada": registro.get("quantidade_informada", ""),
                "estoque_anterior": registro.get("estoque_anterior", ""),
                "estoque_final": registro.get("estoque_final", ""),
                "resultado": registro.get("resultado", ""),
                "motivo": registro.get("motivo", "")
            })

    return caminho_historico, nome_historico


def ler_resumo_historico_importacao_produtos(caminho_arquivo):
    resumo = {
        "data_importacao": "",
        "usuario": "-",
        "arquivo_origem": "-",
        "novos": 0,
        "atualizados": 0,
        "ignorados": 0,
        "total": 0
    }

    with open(caminho_arquivo, "r", newline="", encoding="utf-8-sig") as arquivo_csv:
        reader = csv.DictReader(arquivo_csv, delimiter=";")

        for registro in reader:
            if not resumo["data_importacao"]:
                resumo["data_importacao"] = registro.get("data_importacao", "")
                resumo["usuario"] = registro.get("usuario", "-") or "-"
                resumo["arquivo_origem"] = registro.get("arquivo_origem", "-") or "-"

            resultado = (registro.get("resultado") or "").upper()

            if resultado == "NOVO":
                resumo["novos"] += 1
                resumo["total"] += 1
            elif resultado == "ATUALIZADO":
                resumo["atualizados"] += 1
                resumo["total"] += 1
            elif resultado == "IGNORADO":
                resumo["ignorados"] += 1
                resumo["total"] += 1

    return resumo

@app.route("/produtos/importar", methods=["POST"])
@admin_required
def importar_produtos():
    arquivo = request.files.get("arquivo")

    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo CSV ou Excel.")
        return redirect(url_for("produtos"))

    try:
        linhas = ler_arquivo_planilha(arquivo)
    except Exception as erro:
        flash(f"Erro ao ler arquivo: {erro}")
        return redirect(url_for("produtos"))

    conn = get_db_connection()
    cursor = conn.cursor()

    importados = 0
    atualizados = 0
    registros_ignorados = []
    registros_historico = []

    def registrar_ignorado(
        numero_linha,
        nome,
        categoria,
        marca,
        sku,
        preco_custo,
        preco_venda,
        tipo_atualizacao,
        estoque_atual,
        estoque_minimo,
        motivo,
        estoque_anterior="",
        estoque_final=""
    ):
        registros_ignorados.append({
            "linha": numero_linha,
            "nome": nome,
            "categoria": categoria,
            "marca": marca,
            "sku": sku,
            "preco_custo": preco_custo,
            "preco_venda": preco_venda,
            "tipo_atualizacao": tipo_atualizacao,
            "estoque_atual": estoque_atual,
            "estoque_minimo": estoque_minimo,
            "motivo": motivo
        })
        registros_historico.append({
            "linha": numero_linha,
            "nome": nome,
            "sku": sku,
            "tipo_atualizacao": tipo_atualizacao,
            "quantidade_informada": estoque_atual,
            "estoque_anterior": estoque_anterior,
            "estoque_final": estoque_final,
            "resultado": "IGNORADO",
            "motivo": motivo
        })

    for numero_linha, linha in enumerate(linhas, start=2):
        nome = get_valor(linha, "nome", "produto")
        categoria = get_valor(linha, "categoria")
        marca = get_valor(linha, "marca")
        sku = get_valor(linha, "sku", "codigo", "código")
        preco_custo = converter_float(get_valor(linha, "preco_custo", "preço_custo", "custo"))
        preco_venda = converter_float(get_valor(linha, "preco_venda", "preço_venda", "venda"))
        estoque_atual = converter_int(get_valor(linha, "estoque_atual", "estoque", "quantidade"))
        estoque_minimo = converter_int(get_valor(linha, "estoque_minimo", "estoque_mínimo", "minimo", "mínimo"))
        tipo_atualizacao = get_valor(
            linha,
            "tipo_atualizacao",
            "tipo_atualização",
            "operacao",
            "operação"
        ).strip().upper()

        if not tipo_atualizacao:
            tipo_atualizacao = "SUBSTITUIR"

        motivos = []

        if not nome:
            motivos.append("Nome do produto não informado")

        if not sku:
            motivos.append("SKU/código do produto não informado")

        if tipo_atualizacao not in ("SUBSTITUIR", "INCREMENTAR"):
            motivos.append("Tipo de atualização deve ser SUBSTITUIR ou INCREMENTAR")

        if estoque_atual < 0:
            motivos.append("A quantidade informada não pode ser negativa")

        # Categoria e marca não são obrigatórias na importação.
        if motivos:
            registrar_ignorado(
                numero_linha,
                nome,
                categoria,
                marca,
                sku,
                preco_custo,
                preco_venda,
                tipo_atualizacao,
                estoque_atual,
                estoque_minimo,
                "; ".join(motivos)
            )
            continue

        try:
            cursor.execute(f"SAVEPOINT importacao_produto_{numero_linha}")

            produto_existente = cursor.execute("""
                SELECT id, estoque_atual
                FROM produtos
                WHERE sku = ?
                LIMIT 1
            """, (sku,)).fetchone()

            if produto_existente:
                produto_id = produto_existente["id"]
                estoque_anterior = produto_existente["estoque_atual"]
                estoque_final = (
                    estoque_anterior + estoque_atual
                    if tipo_atualizacao == "INCREMENTAR"
                    else estoque_atual
                )

                cursor.execute("""
                    UPDATE produtos
                    SET
                        nome = ?,
                        categoria = ?,
                        marca = ?,
                        preco_custo = ?,
                        preco_venda = ?,
                        estoque_atual = ?,
                        estoque_minimo = ?,
                        ativo = 1,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (
                    nome,
                    categoria,
                    marca,
                    preco_custo,
                    preco_venda,
                    estoque_final,
                    estoque_minimo,
                    produto_id
                ))

                if estoque_anterior != estoque_final:
                    diferenca = estoque_final - estoque_anterior

                    if tipo_atualizacao == "INCREMENTAR":
                        tipo = "IMPORTACAO_INCREMENTO_ENTRADA"
                        quantidade = estoque_atual
                    elif diferenca > 0:
                        tipo = "IMPORTACAO_AJUSTE_ENTRADA"
                        quantidade = diferenca
                    else:
                        tipo = "IMPORTACAO_AJUSTE_SAIDA"
                        quantidade = abs(diferenca)

                    cursor.execute("""
                        INSERT INTO movimentacoes_estoque (
                            produto_id,
                            tipo,
                            quantidade,
                            estoque_anterior,
                            estoque_atual,
                            observacoes
                        ) VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        produto_id,
                        tipo,
                        quantidade,
                        estoque_anterior,
                        estoque_final,
                        f"{tipo_atualizacao.title()} por importação de planilha"
                    ))

                atualizados += 1
                resultado = "ATUALIZADO"

            else:
                estoque_anterior = 0
                estoque_final = estoque_atual

                cursor.execute("""
                    INSERT INTO produtos (
                        nome,
                        categoria,
                        marca,
                        sku,
                        preco_custo,
                        preco_venda,
                        estoque_atual,
                        estoque_minimo,
                        ativo
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
                """, (
                    nome,
                    categoria,
                    marca,
                    sku,
                    preco_custo,
                    preco_venda,
                    estoque_final,
                    estoque_minimo
                ))

                produto_id = cursor.lastrowid

                cursor.execute("""
                    INSERT INTO movimentacoes_estoque (
                        produto_id,
                        tipo,
                        quantidade,
                        estoque_anterior,
                        estoque_atual,
                        observacoes
                    ) VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    produto_id,
                    "IMPORTACAO_ENTRADA_INICIAL",
                    estoque_final,
                    0,
                    estoque_final,
                    "Estoque inicial por importação de planilha"
                ))

                importados += 1
                resultado = "NOVO"

            cursor.execute(f"RELEASE SAVEPOINT importacao_produto_{numero_linha}")

            registros_historico.append({
                "linha": numero_linha,
                "nome": nome,
                "sku": sku,
                "tipo_atualizacao": tipo_atualizacao,
                "quantidade_informada": estoque_atual,
                "estoque_anterior": estoque_anterior,
                "estoque_final": estoque_final,
                "resultado": resultado,
                "motivo": ""
            })

        except Exception as erro:
            try:
                cursor.execute(f"ROLLBACK TO SAVEPOINT importacao_produto_{numero_linha}")
                cursor.execute(f"RELEASE SAVEPOINT importacao_produto_{numero_linha}")
            except sqlite3.Error:
                pass

            registrar_ignorado(
                numero_linha,
                nome,
                categoria,
                marca,
                sku,
                preco_custo,
                preco_venda,
                tipo_atualizacao,
                estoque_atual,
                estoque_minimo,
                f"Erro ao importar linha: {erro}"
            )
            continue

    conn.commit()
    conn.close()

    caminho_historico, nome_historico = gerar_historico_importacao_produtos(
        secure_filename(arquivo.filename),
        registros_historico
    )

    ignorados = len(registros_ignorados)

    registrar_log(
        "PRODUTOS_IMPORTADOS",
        "produtos",
        descricao=(
            f"Arquivo {secure_filename(arquivo.filename)}. "
            f"Novos: {importados}. Atualizados: {atualizados}. "
            f"Ignorados: {ignorados}. Histórico: {nome_historico}."
        )
    )

    if ignorados > 0:
        caminho_arquivo, nome_arquivo = gerar_planilha_produtos_ignorados(
            registros_ignorados,
            importados,
            atualizados
        )

        return send_file(
            caminho_arquivo,
            as_attachment=True,
            download_name=nome_arquivo
        )

    flash(f"Importação concluída. Novos: {importados}. Atualizados: {atualizados}. Ignorados: {ignorados}.")
    return redirect(url_for("produtos"))


@app.route("/produtos/importacoes")
@admin_required
def historico_importacoes_produtos():
    busca = request.args.get("busca", "").strip().lower()
    page = normalizar_pagina(request.args.get("page", 1))
    per_page = 20
    historico = []

    os.makedirs(HISTORICO_IMPORTACOES_PRODUTOS_DIR, exist_ok=True)

    for nome_arquivo in os.listdir(HISTORICO_IMPORTACOES_PRODUTOS_DIR):
        if not nome_arquivo.lower().endswith(".csv"):
            continue

        caminho_arquivo = os.path.join(
            HISTORICO_IMPORTACOES_PRODUTOS_DIR,
            nome_arquivo
        )

        try:
            resumo = ler_resumo_historico_importacao_produtos(caminho_arquivo)
            resumo["nome_arquivo"] = nome_arquivo

            texto_busca = (
                f"{resumo['arquivo_origem']} {resumo['usuario']} {nome_arquivo}"
            ).lower()

            if not busca or busca in texto_busca:
                historico.append(resumo)
        except (OSError, csv.Error, UnicodeError):
            continue

    historico.sort(
        key=lambda item: (item["data_importacao"], item["nome_arquivo"]),
        reverse=True
    )

    total_registros = len(historico)
    page, total_paginas, offset = calcular_paginacao(
        total_registros,
        page,
        per_page
    )
    historico = historico[offset:offset + per_page]

    return render_template(
        "historico_importacoes_produtos.html",
        historico=historico,
        busca=request.args.get("busca", "").strip(),
        page=page,
        total_paginas=total_paginas,
        total_registros=total_registros,
        usuario=usuario_logado()
    )


@app.route("/produtos/importacoes/<path:nome_arquivo>/baixar")
@admin_required
def baixar_historico_importacao_produtos(nome_arquivo):
    nome_seguro = secure_filename(nome_arquivo)

    if nome_seguro != nome_arquivo or not nome_seguro.lower().endswith(".csv"):
        flash("Arquivo de histórico inválido.")
        return redirect(url_for("historico_importacoes_produtos"))

    caminho_arquivo = os.path.join(
        HISTORICO_IMPORTACOES_PRODUTOS_DIR,
        nome_seguro
    )

    if not os.path.isfile(caminho_arquivo):
        flash("Histórico de importação não encontrado.")
        return redirect(url_for("historico_importacoes_produtos"))

    return send_file(
        caminho_arquivo,
        as_attachment=True,
        download_name=nome_seguro,
        mimetype="text/csv"
    )


@app.route("/produtos/exportar")
@admin_required
def exportar_produtos():
    conn = get_db_connection()

    produtos = conn.execute("""
        SELECT
            nome,
            categoria,
            marca,
            sku,
            preco_custo,
            preco_venda,
            estoque_atual,
            estoque_minimo
        FROM produtos
        WHERE ativo = 1
        ORDER BY nome ASC
    """).fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    ws.append([
        "nome",
        "categoria",
        "marca",
        "sku",
        "preco_custo",
        "preco_venda",
        "estoque_atual",
        "estoque_minimo"
    ])

    for produto in produtos:
        ws.append([
            produto["nome"] or "",
            produto["categoria"] or "",
            produto["marca"] or "",
            produto["sku"] or "",
            produto["preco_custo"] or 0,
            produto["preco_venda"] or 0,
            produto["estoque_atual"] or 0,
            produto["estoque_minimo"] or 0
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 4, 45)

    os.makedirs("exports", exist_ok=True)

    agora = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"produtos_exportados_{agora}.xlsx"
    caminho_arquivo = os.path.join("exports", nome_arquivo)

    wb.save(caminho_arquivo)

    return send_file(
        caminho_arquivo,
        as_attachment=True,
        download_name=nome_arquivo
    )

@app.route("/produtos/modelo")

@admin_required
def baixar_modelo_produtos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Produtos"

    ws.append([
        "nome",
        "categoria",
        "marca",
        "sku",
        "preco_custo",
        "preco_venda",
        "tipo_atualizacao",
        "estoque_atual",
        "estoque_minimo"
    ])

    ws.append([
        "Carregador 20W Tipo-C",
        "Carregadores",
        "Start",
        "CAR-20W-TC",
        35,
        69.90,
        "SUBSTITUIR",
        10,
        3
    ])

    ws.append([
        "Cabo USB-C 1M",
        "Cabos",
        "Start",
        "CAB-USBC-1M",
        8,
        19.90,
        "INCREMENTAR",
        20,
        5
    ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = max_length + 4

    os.makedirs("exports", exist_ok=True)

    caminho_arquivo = os.path.join("exports", "modelo_importacao_produtos.xlsx")
    wb.save(caminho_arquivo)

    return send_file(caminho_arquivo, as_attachment=True)

@app.route("/produtos/<int:produto_id>/editar", methods=["GET", "POST"])
@admin_required
def produto_editar(produto_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    produto = conn.execute("""
        SELECT
            id,
            nome,
            categoria,
            marca,
            sku,
            preco_custo,
            preco_venda,
            estoque_atual,
            estoque_minimo,
            created_at
        FROM produtos
        WHERE id = ? AND ativo = 1
    """, (produto_id,)).fetchone()

    if not produto:
        conn.close()
        flash("Produto não encontrado.")
        return redirect(url_for("produtos"))

    if request.method == "POST":
        nome = request.form.get("nome")
        categoria = request.form.get("categoria")
        marca = request.form.get("marca")
        sku = request.form.get("sku")
        preco_custo = float(request.form.get("preco_custo") or 0)
        preco_venda = float(request.form.get("preco_venda") or 0)
        estoque_minimo = int(request.form.get("estoque_minimo") or 0)

        try:
            conn.execute("""
                UPDATE produtos
                SET
                    nome = ?,
                    categoria = ?,
                    marca = ?,
                    sku = ?,
                    preco_custo = ?,
                    preco_venda = ?,
                    estoque_minimo = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                nome,
                categoria,
                marca,
                sku,
                preco_custo,
                preco_venda,
                estoque_minimo,
                produto_id
            ))

            conn.commit()
            conn.close()

            flash("Produto atualizado com sucesso.")
            return redirect(url_for("produtos"))

        except sqlite3.IntegrityError:
            conn.close()
            flash("Erro: já existe outro produto com este SKU/código.")
            return redirect(url_for("produto_editar", produto_id=produto_id))

    lucro_unitario = float(produto["preco_venda"] or 0) - float(produto["preco_custo"] or 0)
    margem_produto = (
        lucro_unitario / float(produto["preco_custo"] or 0) * 100
        if float(produto["preco_custo"] or 0) > 0
        else 0
    )
    valor_estoque = float(produto["estoque_atual"] or 0) * float(produto["preco_custo"] or 0)

    conn.close()

    return render_template(
        "produto_editar.html",
        produto=produto,
        lucro_unitario=lucro_unitario,
        margem_produto=margem_produto,
        valor_estoque=valor_estoque
    )

@app.route("/produtos/<int:produto_id>/inativar", methods=["POST"])
@admin_required
def produto_inativar(produto_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    produto = conn.execute("""
        SELECT id
        FROM produtos
        WHERE id = ? AND ativo = 1
    """, (produto_id,)).fetchone()

    if not produto:
        conn.close()
        flash("Produto não encontrado.")
        return redirect(url_for("produtos"))

    conn.execute("""
        UPDATE produtos
        SET ativo = 0,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (produto_id,))

    conn.commit()
    conn.close()

    registrar_log(
        "PRODUTO_INATIVADO",
        "produtos",
        produto_id,
        f"Produto #{produto_id} foi inativado."
    )

    flash("Produto inativado com sucesso.")
    return redirect(url_for("produtos"))

@app.route("/produtos/<int:produto_id>/reativar", methods=["POST"])
@admin_required
def produto_reativar(produto_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    produto = conn.execute("""
        SELECT id
        FROM produtos
        WHERE id = ? AND ativo = 0
    """, (produto_id,)).fetchone()

    if not produto:
        conn.close()
        flash("Produto inativo não encontrado.")
        return redirect(url_for("produtos", status="inativos"))

    conn.execute("""
        UPDATE produtos
        SET ativo = 1,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (produto_id,))

    conn.commit()
    conn.close()

    registrar_log(
        "PRODUTO_REATIVADO",
        "produtos",
        produto_id,
        f"Produto #{produto_id} foi reativado."
    )

    flash("Produto reativado com sucesso.")
    return redirect(url_for("produtos", status="ativos"))

@app.route("/caixa", methods=["GET", "POST"])
@login_required
def caixa():
    conn = get_db_connection()

    caixa_aberto = obter_caixa_aberto(conn)

    resumo_caixa = None
    movimentacoes = []
    vendas_caixa = []
    historico_caixas = []

    if request.method == "POST":
        acao = request.form.get("acao")

        if acao == "abrir":
            if caixa_aberto:
                conn.close()
                flash("Já existe um caixa aberto.")
                return redirect(url_for("caixa"))

            valor_inicial = converter_float(request.form.get("valor_inicial"))
            observacoes = request.form.get("observacoes_abertura", "").strip()
            usuario = usuario_logado()

            cursor = conn.cursor()

            cursor.execute("""
                INSERT INTO caixas (
                    usuario_abertura_id,
                    usuario_abertura_nome,
                    valor_inicial,
                    observacoes_abertura,
                    status
                ) VALUES (?, ?, ?, ?, 'ABERTO')
            """, (
                usuario["id"],
                usuario["nome"],
                valor_inicial,
                observacoes
            ))

            caixa_id = cursor.lastrowid

            conn.commit()
            conn.close()

            registrar_log(
                "CAIXA_ABERTO",
                "caixas",
                caixa_id,
                f"Caixa #{caixa_id} aberto com valor inicial R$ {valor_inicial:.2f}."
            )

            flash("Caixa aberto com sucesso. Você já pode registrar vendas.")
            return redirect(url_for("vendas"))

        if acao == "movimentar":
            if not caixa_aberto:
                conn.close()
                flash("Não existe caixa aberto.")
                return redirect(url_for("caixa"))

            tipo = request.form.get("tipo_movimento")
            valor = converter_float(request.form.get("valor_movimento"))
            descricao = request.form.get("descricao_movimento", "").strip()
            usuario = usuario_logado()

            if tipo not in ["ENTRADA", "SAIDA"]:
                conn.close()
                flash("Tipo de movimentação inválido.")
                return redirect(url_for("caixa"))

            if valor <= 0:
                conn.close()
                flash("Informe um valor maior que zero.")
                return redirect(url_for("caixa"))

            conn.execute("""
                INSERT INTO caixa_movimentacoes (
                    caixa_id,
                    tipo,
                    valor,
                    descricao,
                    usuario_id,
                    usuario_nome
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                caixa_aberto["id"],
                tipo,
                valor,
                descricao,
                usuario["id"],
                usuario["nome"]
            ))

            conn.commit()
            conn.close()

            registrar_log(
                "CAIXA_MOVIMENTACAO",
                "caixas",
                caixa_aberto["id"],
                f"Movimentação de caixa: {tipo} R$ {valor:.2f}. {descricao}"
            )

            flash("Movimentação registrada com sucesso.")
            return redirect(url_for("caixa"))

        if acao == "fechar":
            if not caixa_aberto:
                conn.close()
                flash("Não existe caixa aberto para fechar.")
                return redirect(url_for("caixa"))

            valor_informado = converter_float(request.form.get("valor_informado"))
            observacoes_fechamento = request.form.get("observacoes_fechamento", "").strip()
            usuario = usuario_logado()

            resumo = calcular_resumo_caixa(conn, caixa_aberto["id"])
            valor_esperado = resumo["valor_esperado"]
            diferenca = valor_informado - valor_esperado

            conn.execute("""
                UPDATE caixas
                SET
                    usuario_fechamento_id = ?,
                    usuario_fechamento_nome = ?,
                    data_fechamento = CURRENT_TIMESTAMP,
                    total_vendas = ?,
                    total_dinheiro = ?,
                    total_pix = ?,
                    total_cartao = ?,
                    total_outros = ?,
                    entradas_manuais = ?,
                    saidas_manuais = ?,
                    valor_esperado = ?,
                    valor_informado = ?,
                    diferenca = ?,
                    observacoes_fechamento = ?,
                    status = 'FECHADO'
                WHERE id = ?
            """, (
                usuario["id"],
                usuario["nome"],
                resumo["total_vendas"],
                resumo["total_dinheiro"],
                resumo["total_pix"],
                resumo["total_cartao"],
                resumo["total_outros"],
                resumo["entradas_manuais"],
                resumo["saidas_manuais"],
                valor_esperado,
                valor_informado,
                diferenca,
                observacoes_fechamento,
                caixa_aberto["id"]
            ))

            conn.commit()
            conn.close()

            registrar_log(
                "CAIXA_FECHADO",
                "caixas",
                caixa_aberto["id"],
                f"Caixa #{caixa_aberto['id']} fechado. Esperado R$ {valor_esperado:.2f}, informado R$ {valor_informado:.2f}, diferença R$ {diferenca:.2f}."
            )

            flash("Caixa fechado com sucesso.")
            return redirect(url_for("caixa"))

    caixa_aberto = obter_caixa_aberto(conn)

    if caixa_aberto:
        resumo_caixa = calcular_resumo_caixa(conn, caixa_aberto["id"])

        movimentacoes = conn.execute("""
            SELECT
                tipo,
                valor,
                descricao,
                usuario_nome,
                created_at
            FROM caixa_movimentacoes
            WHERE caixa_id = ?
            ORDER BY id DESC
            LIMIT 20
        """, (
            caixa_aberto["id"],
        )).fetchall()

        vendas_caixa = conn.execute("""
            SELECT
                id,
                data_venda,
                cliente_id,
                vendedor,
                forma_pagamento,
                valor_total,
                status
            FROM vendas
            WHERE caixa_id = ?
            ORDER BY id DESC
            LIMIT 20
        """, (
            caixa_aberto["id"],
        )).fetchall()

    usuario_atual = usuario_logado()

    if usuario_atual["perfil"] == "ADMIN":
        historico_caixas = conn.execute("""
            SELECT
                id,
                data_abertura,
                data_fechamento,
                usuario_abertura_nome,
                usuario_fechamento_nome,
                valor_inicial,
                valor_esperado,
                valor_informado,
                diferenca,
                status
            FROM caixas
            ORDER BY id DESC
            LIMIT 20
        """).fetchall()
    else:
        historico_caixas = conn.execute("""
            SELECT
                id,
                data_abertura,
                data_fechamento,
                usuario_abertura_nome,
                usuario_fechamento_nome,
                valor_inicial,
                valor_esperado,
                valor_informado,
                diferenca,
                status
            FROM caixas
            WHERE usuario_abertura_id = ?
               OR status = 'ABERTO'
            ORDER BY id DESC
            LIMIT 20
        """, (
            usuario_atual["id"],
        )).fetchall()

    conn.close()

    return render_template(
        "caixa.html",
        caixa_aberto=caixa_aberto,
        resumo_caixa=resumo_caixa,
        movimentacoes=movimentacoes,
        vendas_caixa=vendas_caixa,
        historico_caixas=historico_caixas,
        usuario=usuario_logado()
    )

@app.route("/caixa/<int:caixa_id>")
@login_required
def caixa_detalhe(caixa_id):
    conn = get_db_connection()

    caixa = conn.execute("""
        SELECT
            id,
            usuario_abertura_id,
            usuario_abertura_nome,
            usuario_fechamento_id,
            usuario_fechamento_nome,
            data_abertura,
            data_fechamento,
            valor_inicial,
            total_vendas,
            total_dinheiro,
            total_pix,
            total_cartao,
            total_outros,
            entradas_manuais,
            saidas_manuais,
            valor_esperado,
            valor_informado,
            diferenca,
            observacoes_abertura,
            observacoes_fechamento,
            status
        FROM caixas
        WHERE id = ?
    """, (caixa_id,)).fetchone()

    if not caixa:
        conn.close()
        flash("Caixa não encontrado.")
        return redirect(url_for("caixa"))

    usuario_atual = usuario_logado()

    if usuario_atual["perfil"] != "ADMIN" and caixa["usuario_abertura_id"] != usuario_atual["id"]:
        conn.close()
        flash("Você não tem permissão para visualizar este caixa.")
        return redirect(url_for("caixa"))

    resumo_atual = None

    if caixa["status"] == "ABERTO":
        resumo_atual = calcular_resumo_caixa(conn, caixa_id)

    vendas_caixa = conn.execute("""
        SELECT
            v.id,
            v.data_venda,
            COALESCE(c.nome, 'Cliente não identificado') AS cliente_nome,
            v.vendedor,
            v.forma_pagamento,
            v.valor_total,
            v.lucro_total,
            v.status
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE v.caixa_id = ?
        ORDER BY v.id DESC
    """, (caixa_id,)).fetchall()

    movimentacoes = conn.execute("""
        SELECT
            id,
            tipo,
            valor,
            descricao,
            usuario_nome,
            created_at
        FROM caixa_movimentacoes
        WHERE caixa_id = ?
        ORDER BY id DESC
    """, (caixa_id,)).fetchall()

    conn.close()

    return render_template(
        "caixa_detalhe.html",
        caixa=caixa,
        resumo_atual=resumo_atual,
        vendas_caixa=vendas_caixa,
        movimentacoes=movimentacoes,
        usuario=usuario_logado()
    )

@app.route("/caixa/<int:caixa_id>/imprimir")
@login_required
def caixa_imprimir(caixa_id):
    conn = get_db_connection()

    caixa = conn.execute("""
        SELECT
            id,
            usuario_abertura_id,
            usuario_abertura_nome,
            usuario_fechamento_nome,
            data_abertura,
            data_fechamento,
            valor_inicial,
            total_vendas,
            total_dinheiro,
            total_pix,
            total_cartao,
            total_outros,
            entradas_manuais,
            saidas_manuais,
            valor_esperado,
            valor_informado,
            diferenca,
            observacoes_abertura,
            observacoes_fechamento,
            status
        FROM caixas
        WHERE id = ?
    """, (caixa_id,)).fetchone()

    if not caixa:
        conn.close()
        flash("Caixa não encontrado.")
        return redirect(url_for("caixa"))

    usuario_atual = usuario_logado()

    if usuario_atual["perfil"] != "ADMIN" and caixa["usuario_abertura_id"] != usuario_atual["id"]:
        conn.close()
        flash("Você não tem permissão para imprimir este caixa.")
        return redirect(url_for("caixa"))

    resumo_atual = None

    if caixa["status"] == "ABERTO":
        resumo_atual = calcular_resumo_caixa(conn, caixa_id)

    vendas_caixa = conn.execute("""
        SELECT
            v.id,
            v.data_venda,
            COALESCE(c.nome, 'Cliente não identificado') AS cliente_nome,
            v.vendedor,
            v.forma_pagamento,
            v.valor_total,
            v.status
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE v.caixa_id = ?
        ORDER BY v.id ASC
    """, (caixa_id,)).fetchall()

    movimentacoes = conn.execute("""
        SELECT
            tipo,
            valor,
            descricao,
            usuario_nome,
            created_at
        FROM caixa_movimentacoes
        WHERE caixa_id = ?
        ORDER BY id ASC
    """, (caixa_id,)).fetchall()

    config = conn.execute("""
        SELECT
            nome_loja,
            telefone,
            endereco,
            cidade,
            instagram
        FROM configuracoes_loja
        WHERE id = 1
    """).fetchone()

    conn.close()

    return render_template(
        "caixa_imprimir.html",
        caixa=caixa,
        resumo_atual=resumo_atual,
        vendas_caixa=vendas_caixa,
        movimentacoes=movimentacoes,
        config=config
    )

@app.route("/vendas", methods=["GET", "POST"])
@login_required
def vendas():
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    caixa_aberto = obter_caixa_aberto(conn)

    if request.method == "POST" and not caixa_aberto:
        conn.close()
        flash("Abra o caixa antes de registrar uma venda.")
        return redirect(url_for("vendas"))

    if request.method == "POST":
        cliente_id = request.form.get("cliente_id")
        vendedor_id = request.form.get("vendedor_id")
        forma_pagamento = request.form.get("forma_pagamento")
        desconto_total = float(request.form.get("desconto_total") or 0)
        observacoes = request.form.get("observacoes")
        itens_json = request.form.get("itens_json")
        valor_final = converter_float(
            request.form.get("valor_final") or request.form.get("valor_final_input")
        )

        pagamento_dividido = request.form.get("pagamento_dividido") == "1"

        pagamentos = []

        if pagamento_dividido:
            valor_pix = converter_float(request.form.get("valor_pix"))
            valor_dinheiro = converter_float(request.form.get("valor_dinheiro"))
            valor_cartao = converter_float(request.form.get("valor_cartao"))
            valor_outros = converter_float(request.form.get("valor_outros"))

            if valor_pix > 0:
                pagamentos.append(("PIX", valor_pix))

            if valor_dinheiro > 0:
                pagamentos.append(("DINHEIRO", valor_dinheiro))

            if valor_cartao > 0:
                pagamentos.append(("CARTAO", valor_cartao))

            if valor_outros > 0:
                pagamentos.append(("OUTROS", valor_outros))

            total_pagamentos = sum(valor for _, valor in pagamentos)

            if abs(total_pagamentos - valor_final) > 0.009:
                conn.close()
                flash("O total dos pagamentos precisa ser igual ao valor final da venda.")
                return redirect(url_for("vendas"))

            forma_pagamento = "MULTIPLO"

        else:
            forma_pagamento = request.form.get("forma_pagamento", "").strip()

            if not forma_pagamento:
                conn.close()
                flash("Informe a forma de pagamento.")
                return redirect(url_for("vendas"))

            pagamentos.append((forma_pagamento, valor_final))

        cliente_id = int(cliente_id) if cliente_id else None

        if cliente_id:
            cliente = conn.execute("""
                SELECT id
                FROM clientes
                WHERE id = ? AND ativo = 1
            """, (cliente_id,)).fetchone()

            if not cliente:
                flash("Cliente selecionado não encontrado ou inativo.")
                conn.close()
                return redirect(url_for("vendas"))

        if not vendedor_id:
            flash("Selecione o vendedor responsável pela venda.")
            conn.close()
            return redirect(url_for("vendas"))

        if not forma_pagamento:
            flash("Selecione a forma de pagamento.")
            conn.close()
            return redirect(url_for("vendas"))
        
        vendedor_usuario = conn.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE id = ?
            AND perfil = 'VENDEDOR'
            AND ativo = 1
        """, (vendedor_id,)).fetchone()

        if not vendedor_usuario:
            flash("Vendedor não encontrado ou inativo.")
            conn.close()
            return redirect(url_for("vendas"))

        vendedor = vendedor_usuario["nome"]

        try:
            itens = json.loads(itens_json) if itens_json else []
        except json.JSONDecodeError:
            flash("Erro ao ler os itens da venda.")
            conn.close()
            return redirect(url_for("vendas"))

        if not itens:
            flash("Adicione pelo menos um produto à venda.")
            conn.close()
            return redirect(url_for("vendas"))

        cursor = conn.cursor()

        produtos_validados = []
        subtotal_bruto = 0
        custo_total = 0

        for item in itens:
            produto_id = int(item.get("produto_id"))
            quantidade = int(item.get("quantidade") or 0)

            if quantidade <= 0:
                flash("A quantidade dos produtos deve ser maior que zero.")
                conn.close()
                return redirect(url_for("vendas"))

            produto = conn.execute("""
                SELECT id, nome, preco_custo, preco_venda, estoque_atual
                FROM produtos
                WHERE id = ? AND ativo = 1
            """, (produto_id,)).fetchone()

            if not produto:
                flash("Um dos produtos não foi encontrado.")
                conn.close()
                return redirect(url_for("vendas"))

            if quantidade > produto["estoque_atual"]:
                flash(f"Estoque insuficiente para o produto: {produto['nome']}.")
                conn.close()
                return redirect(url_for("vendas"))

            preco_unitario = float(produto["preco_venda"])
            preco_custo_unitario = float(produto["preco_custo"])

            subtotal_item = preco_unitario * quantidade
            custo_item = preco_custo_unitario * quantidade

            subtotal_bruto += subtotal_item
            custo_total += custo_item

            produtos_validados.append({
                "produto": produto,
                "quantidade": quantidade,
                "preco_unitario": preco_unitario,
                "preco_custo_unitario": preco_custo_unitario,
                "subtotal_item": subtotal_item,
                "custo_item": custo_item
            })

        valor_total = subtotal_bruto - desconto_total
        lucro_total = valor_total - custo_total

        if valor_total < 0:
            flash("O desconto não pode ser maior que o valor da venda.")
            conn.close()
            return redirect(url_for("vendas"))
                
        caixa_aberto = obter_caixa_aberto(conn)

        if not caixa_aberto:
            flash("Abra o caixa antes de registrar uma venda.")
            conn.close()
            return redirect(url_for("vendas"))

        cursor.execute("""
            INSERT INTO vendas (
            cliente_id,
            vendedor,
            vendedor_id,
            forma_pagamento,
            desconto_total,
            valor_total,
            custo_total,
            lucro_total,
            observacoes,
            caixa_id
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            cliente_id,
            vendedor,
            vendedor_id,
            forma_pagamento,
            desconto_total,
            valor_total,
            custo_total,
            lucro_total,
            observacoes,
            caixa_aberto["id"]
        ))

        venda_id = cursor.lastrowid

        for forma, valor in pagamentos:
            cursor.execute("""
                INSERT INTO venda_pagamentos (
                    venda_id,
                    forma_pagamento,
                    valor
                ) VALUES (?, ?, ?)
            """, (
                venda_id,
                forma,
                valor
            ))

        for item in produtos_validados:
            produto = item["produto"]
            quantidade = item["quantidade"]
            preco_unitario = item["preco_unitario"]
            preco_custo_unitario = item["preco_custo_unitario"]
            subtotal_item = item["subtotal_item"]
            custo_item = item["custo_item"]
            lucro_item = subtotal_item - custo_item

            cursor.execute("""
                INSERT INTO venda_itens (
                    venda_id,
                    produto_id,
                    quantidade,
                    preco_unitario,
                    preco_custo_unitario,
                    desconto,
                    subtotal,
                    lucro_item
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                venda_id,
                produto["id"],
                quantidade,
                preco_unitario,
                preco_custo_unitario,
                0,
                subtotal_item,
                lucro_item
            ))

            estoque_anterior = produto["estoque_atual"]
            estoque_atual = estoque_anterior - quantidade

            cursor.execute("""
                UPDATE produtos
                SET estoque_atual = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                estoque_atual,
                produto["id"]
            ))

            cursor.execute("""
                INSERT INTO movimentacoes_estoque (
                    produto_id,
                    tipo,
                    quantidade,
                    estoque_anterior,
                    estoque_atual,
                    referencia_venda_id,
                    observacoes
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                produto["id"],
                "SAIDA_VENDA",
                quantidade,
                estoque_anterior,
                estoque_atual,
                venda_id,
                f"Venda #{venda_id}"
            ))

        conn.commit()
        cliente_log = cliente_id if cliente_id else "não identificado"

        registrar_log(
            "VENDA_CRIADA",
            "vendas",
            venda_id,
            f"Venda #{venda_id} criada para o cliente {cliente_log}, vendedor {vendedor}, total R$ {valor_total:.2f}."
        )
        conn.close()

        return redirect(url_for(
            "venda_recibo_termico",
            venda_id=venda_id,
            auto=1
        ))

    vendedores = conn.execute("""
        SELECT id, nome, email
        FROM usuarios
        WHERE perfil = 'VENDEDOR'
        AND ativo = 1
        ORDER BY nome ASC
    """).fetchall()

    busca = request.args.get("busca", "").strip()
    status = request.args.get("status", "todas")
    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip()

    page = int(request.args.get("page", 1))
    per_page = 10

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    where_clauses = ["1 = 1"]
    params = []

    if busca:
        termo = f"%{busca}%"

        where_clauses.append("""
            (
                v.id LIKE ?
                OR c.nome LIKE ?
                OR c.telefone LIKE ?
                OR v.vendedor LIKE ?
                OR v.forma_pagamento LIKE ?
            )
        """)

        params.extend([termo, termo, termo, termo, termo])

    if status == "concluidas":
        where_clauses.append("v.status = 'CONCLUIDA'")
    elif status == "canceladas":
        where_clauses.append("v.status = 'CANCELADA'")

    if data_inicio:
        where_clauses.append("date(v.data_venda) >= date(?)")
        params.append(data_inicio)

    if data_fim:
        where_clauses.append("date(v.data_venda) <= date(?)")
        params.append(data_fim)

    where_sql = " AND ".join(where_clauses)

    total_registros = conn.execute(f"""
        SELECT COUNT(*) AS total
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE {where_sql}
    """, params).fetchone()["total"]

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas == 0:
        total_paginas = 1

    if page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    vendas_lista = conn.execute(f"""
        SELECT 
            v.id,
            v.data_venda,
            COALESCE(c.nome, 'Cliente não identificado') AS cliente_nome,
            v.vendedor,
            v.forma_pagamento,
            v.valor_total,
            v.custo_total,
            v.lucro_total,
            v.status
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE {where_sql}
        ORDER BY v.id DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()

    conn.close()

    return render_template(
        "vendas.html",
        vendedores=vendedores,
        vendas=vendas_lista,
        busca=busca,
        status=status,
        data_inicio=data_inicio,
        data_fim=data_fim,
        page=page,
        total_paginas=total_paginas,
        total_registros=total_registros,
        caixa_aberto=caixa_aberto,
        pdv_bloqueado=not caixa_aberto,
        usuario=usuario_logado()
    )

@app.route("/api/clientes/busca")
def api_buscar_clientes():
    if "usuario_id" not in session:
        return jsonify({"erro": "Não autorizado"}), 401

    termo = request.args.get("q", "").strip()

    if len(termo) < 3:
        return jsonify([])

    termo_like = f"%{termo}%"

    conn = get_db_connection()

    clientes = conn.execute("""
        SELECT
            id,
            nome,
            telefone,
            endereco_completo
        FROM clientes
        WHERE ativo = 1
          AND (
              nome LIKE ?
              OR telefone LIKE ?
              OR endereco_completo LIKE ?
          )
        ORDER BY nome ASC
        LIMIT 20
    """, (
        termo_like,
        termo_like,
        termo_like
    )).fetchall()

    conn.close()

    resultado = []

    for cliente in clientes:
        resultado.append({
            "id": cliente["id"],
            "nome": cliente["nome"],
            "telefone": cliente["telefone"],
            "endereco_completo": cliente["endereco_completo"]
        })

    return jsonify(resultado)

@app.route("/api/clientes/criar", methods=["POST"])
@login_required
def api_criar_cliente():
    dados = request.get_json() or {}

    nome = (dados.get("nome") or "").strip()
    telefone = (dados.get("telefone") or "").strip()
    endereco_completo = (dados.get("endereco_completo") or "").strip()
    observacoes = (dados.get("observacoes") or "").strip()
    tags_texto = (dados.get("tags") or "").strip()

    if not nome:
        return jsonify({
            "sucesso": False,
            "mensagem": "Informe o nome do cliente."
        }), 400

    if not telefone:
        return jsonify({
            "sucesso": False,
            "mensagem": "Informe o telefone do cliente."
        }), 400

    telefone_normalizado = "".join(c for c in telefone if c.isdigit())

    conn = get_db_connection()
    cursor = conn.cursor()

    cliente_existente = cursor.execute("""
        SELECT
            id,
            nome,
            telefone
        FROM clientes
        WHERE REPLACE(REPLACE(REPLACE(REPLACE(telefone, ' ', ''), '-', ''), '(', ''), ')', '') = ?
        LIMIT 1
    """, (
        telefone_normalizado,
    )).fetchone()

    if cliente_existente:
        conn.close()

        return jsonify({
            "sucesso": False,
            "mensagem": f"Já existe um cliente cadastrado com este telefone: {cliente_existente['nome']}."
        }), 409

    cursor.execute("""
        INSERT INTO clientes (
            nome,
            telefone,
            endereco_completo,
            observacoes,
            ativo
        ) VALUES (?, ?, ?, ?, 1)
    """, (
        nome,
        telefone_normalizado,
        endereco_completo,
        observacoes
    ))

    cliente_id = cursor.lastrowid

    if tags_texto:
        tags = [tag.strip() for tag in tags_texto.split(",") if tag.strip()]

        for tag in tags:
            cursor.execute("""
                INSERT INTO cliente_tags (
                    cliente_id,
                    tag
                ) VALUES (?, ?)
            """, (
                cliente_id,
                tag
            ))

    conn.commit()

    cliente_criado = conn.execute("""
        SELECT
            id,
            nome,
            telefone,
            endereco_completo
        FROM clientes
        WHERE id = ?
    """, (
        cliente_id,
    )).fetchone()

    conn.close()

    registrar_log(
        "CLIENTE_CRIADO_VENDA",
        "clientes",
        cliente_id,
        f"Cliente criado rapidamente pela tela de vendas: {nome}."
    )

    return jsonify({
        "sucesso": True,
        "mensagem": "Cliente criado com sucesso.",
        "cliente": {
            "id": cliente_criado["id"],
            "nome": cliente_criado["nome"],
            "telefone": cliente_criado["telefone"],
            "endereco_completo": cliente_criado["endereco_completo"] or ""
        }
    })

@app.route("/api/produtos/busca")
def api_buscar_produtos():
    if "usuario_id" not in session:
        return jsonify({"erro": "Não autorizado"}), 401

    termo = request.args.get("q", "").strip()

    if len(termo) < 3:
        return jsonify([])

    termo_like = f"%{termo}%"

    conn = get_db_connection()

    produtos = conn.execute("""
        SELECT
            id,
            nome,
            sku,
            categoria,
            marca,
            preco_venda,
            estoque_atual,
            CASE 
                WHEN sku = ? THEN 0
                WHEN sku LIKE ? THEN 1
                WHEN nome LIKE ? THEN 2
                ELSE 3
            END AS ordem
        FROM produtos
        WHERE ativo = 1
        AND estoque_atual > 0
        AND (
            sku = ?
            OR sku LIKE ?
            OR nome LIKE ?
            OR categoria LIKE ?
            OR marca LIKE ?
        )
        ORDER BY ordem ASC, nome ASC
        LIMIT 20
    """, (
        termo,
        termo_like,
        termo_like,
        termo,
        termo_like,
        termo_like,
        termo_like,
        termo_like
    )).fetchall()

    conn.close()

    resultado = []

    for produto in produtos:
        resultado.append({
            "id": produto["id"],
            "nome": produto["nome"],
            "sku": produto["sku"],
            "categoria": produto["categoria"],
            "marca": produto["marca"],
            "preco_venda": produto["preco_venda"],
            "estoque_atual": produto["estoque_atual"]
        })

    return jsonify(resultado)

@app.route("/api/produtos/criar", methods=["POST"])
@login_required
def api_criar_produto():
    dados = request.get_json() or {}

    nome = (dados.get("nome") or "").strip()
    categoria = (dados.get("categoria") or "").strip()
    marca = (dados.get("marca") or "").strip()
    sku = (dados.get("sku") or "").strip()
    preco_custo = converter_float(dados.get("preco_custo"))
    preco_venda = converter_float(dados.get("preco_venda"))
    estoque_atual = int(converter_float(dados.get("estoque_atual")))
    estoque_minimo = int(converter_float(dados.get("estoque_minimo")))

    if not nome:
        return jsonify({
            "sucesso": False,
            "mensagem": "Informe o nome do produto."
        }), 400

    if not sku:
        return jsonify({
            "sucesso": False,
            "mensagem": "Informe o SKU/código do produto."
        }), 400

    if preco_venda <= 0:
        return jsonify({
            "sucesso": False,
            "mensagem": "Informe um preço de venda maior que zero."
        }), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    produto_existente = cursor.execute("""
        SELECT
            id,
            nome,
            sku,
            preco_venda,
            estoque_atual
        FROM produtos
        WHERE sku = ?
        LIMIT 1
    """, (sku,)).fetchone()

    if produto_existente:
        conn.close()

        return jsonify({
            "sucesso": False,
            "mensagem": f"Já existe um produto cadastrado com este SKU: {produto_existente['nome']}."
        }), 409

    cursor.execute("""
        INSERT INTO produtos (
            nome,
            categoria,
            marca,
            sku,
            preco_custo,
            preco_venda,
            estoque_atual,
            estoque_minimo,
            ativo
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
    """, (
        nome,
        categoria,
        marca,
        sku,
        preco_custo,
        preco_venda,
        estoque_atual,
        estoque_minimo
    ))

    produto_id = cursor.lastrowid

    if estoque_atual > 0:
        cursor.execute("""
            INSERT INTO movimentacoes_estoque (
                produto_id,
                tipo,
                quantidade,
                estoque_anterior,
                estoque_atual,
                observacoes
            ) VALUES (?, ?, ?, ?, ?, ?)
        """, (
            produto_id,
            "ENTRADA_INICIAL",
            estoque_atual,
            0,
            estoque_atual,
            "Cadastro rápido pela tela de vendas"
        ))

    conn.commit()

    produto_criado = conn.execute("""
        SELECT
            id,
            nome,
            sku,
            categoria,
            marca,
            preco_custo,
            preco_venda,
            estoque_atual
        FROM produtos
        WHERE id = ?
    """, (produto_id,)).fetchone()

    conn.close()

    registrar_log(
        "PRODUTO_CRIADO_VENDA",
        "produtos",
        produto_id,
        f"Produto criado rapidamente pela tela de vendas: {nome}."
    )

    return jsonify({
        "sucesso": True,
        "mensagem": "Produto criado com sucesso.",
        "produto": {
            "id": produto_criado["id"],
            "nome": produto_criado["nome"],
            "sku": produto_criado["sku"],
            "categoria": produto_criado["categoria"] or "",
            "marca": produto_criado["marca"] or "",
            "preco_custo": produto_criado["preco_custo"] or 0,
            "preco_venda": produto_criado["preco_venda"] or 0,
            "estoque_atual": produto_criado["estoque_atual"] or 0
        }
    })

@app.route("/vendas/<int:venda_id>")
def venda_detalhe(venda_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    venda = conn.execute("""
        SELECT
            v.id,
            v.data_venda,
            v.vendedor,
            v.vendedor_id,
            v.forma_pagamento,
            v.desconto_total,
            v.valor_total,
            v.custo_total,
            v.lucro_total,
            v.observacoes,
            v.status,
            v.motivo_cancelamento,
            v.data_cancelamento,
            v.cancelado_por,
            c.id AS cliente_id,
            COALESCE(c.nome, 'Cliente não identificado') AS cliente_nome,
            COALESCE(c.telefone, '-') AS cliente_telefone,
            COALESCE(c.endereco_completo, '-') AS cliente_endereco
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE v.id = ?
    """, (venda_id,)).fetchone()

    if not venda:
        conn.close()
        flash("Venda não encontrada.")
        return redirect(url_for("vendas"))

    itens = conn.execute("""
        SELECT
            vi.id,
            vi.quantidade,
            vi.preco_unitario,
            vi.preco_custo_unitario,
            vi.desconto,
            vi.subtotal,
            vi.lucro_item,
            p.nome AS produto_nome,
            p.sku,
            p.categoria
        FROM venda_itens vi
        INNER JOIN produtos p ON p.id = vi.produto_id
        WHERE vi.venda_id = ?
        ORDER BY vi.id ASC
    """, (venda_id,)).fetchall()

    movimentacoes = conn.execute("""
        SELECT
            m.tipo,
            m.quantidade,
            m.estoque_anterior,
            m.estoque_atual,
            m.observacoes,
            m.created_at,
            p.nome AS produto_nome,
            p.sku AS produto_sku
        FROM movimentacoes_estoque m
        INNER JOIN produtos p ON p.id = m.produto_id
        WHERE m.referencia_venda_id = ?
        ORDER BY m.id ASC
    """, (venda_id,)).fetchall()

    pagamentos = conn.execute("""
        SELECT
            forma_pagamento,
            valor
        FROM venda_pagamentos
        WHERE venda_id = ?
        ORDER BY id
    """, (venda_id,)).fetchall()

    conn.close()

    return render_template(
        "venda_detalhe.html",
        venda=venda,
        itens=itens,
        movimentacoes=movimentacoes,
        pagamentos=pagamentos,
        usuario=usuario_logado()
    )

@app.route("/vendas/<int:venda_id>/recibo")
def venda_recibo(venda_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    venda = conn.execute("""
        SELECT
            v.id,
            v.data_venda,
            v.vendedor,
            v.forma_pagamento,
            v.desconto_total,
            v.valor_total,
            v.custo_total,
            v.lucro_total,
            v.observacoes,
            v.status,
            COALESCE(c.nome, 'Cliente não identificado') AS cliente_nome,
            COALESCE(c.telefone, '-') AS cliente_telefone,
            COALESCE(c.endereco_completo, '-') AS cliente_endereco
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE v.id = ?
    """, (venda_id,)).fetchone()

    if not venda:
        conn.close()
        flash("Venda não encontrada.")
        return redirect(url_for("vendas"))

    itens = conn.execute("""
        SELECT
            vi.quantidade,
            vi.preco_unitario,
            vi.subtotal,
            p.nome AS produto_nome,
            p.sku
        FROM venda_itens vi
        INNER JOIN produtos p ON p.id = vi.produto_id
        WHERE vi.venda_id = ?
        ORDER BY vi.id ASC
    """, (venda_id,)).fetchall()

    config = conn.execute("""
        SELECT
            nome_loja,
            telefone,
            endereco,
            cidade,
            instagram,
            mensagem_recibo
        FROM configuracoes_loja
        WHERE id = 1
    """).fetchone()

    conn.close()

    return render_template(
        "recibo.html",
        venda=venda,
        itens=itens,
        config=config
    )

@app.route("/vendas/<int:venda_id>/recibo-termico")
def venda_recibo_termico(venda_id):
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    venda = conn.execute("""
        SELECT
            v.id,
            v.data_venda,
            v.vendedor,
            v.forma_pagamento,
            v.desconto_total,
            v.valor_total,
            v.observacoes,
            v.status,
            COALESCE(c.nome, 'Cliente não identificado') AS cliente_nome,
            COALESCE(c.telefone, '-') AS cliente_telefone,
            COALESCE(c.endereco_completo, '-') AS cliente_endereco
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE v.id = ?
    """, (venda_id,)).fetchone()

    if not venda:
        conn.close()
        flash("Venda não encontrada.")
        return redirect(url_for("vendas"))

    itens = conn.execute("""
        SELECT
            vi.quantidade,
            vi.preco_unitario,
            vi.subtotal,
            p.nome AS produto_nome,
            p.sku
        FROM venda_itens vi
        INNER JOIN produtos p ON p.id = vi.produto_id
        WHERE vi.venda_id = ?
        ORDER BY vi.id ASC
    """, (venda_id,)).fetchall()

    pagamentos = conn.execute("""
        SELECT
            forma_pagamento,
            valor
        FROM venda_pagamentos
        WHERE venda_id = ?
        ORDER BY id ASC
    """, (venda_id,)).fetchall()

    config = conn.execute("""
        SELECT
            nome_loja,
            telefone,
            endereco,
            cidade,
            instagram,
            mensagem_recibo
        FROM configuracoes_loja
        WHERE id = 1
    """).fetchone()

    total_itens = sum(float(item["subtotal"] or 0) for item in itens)
    quantidade_itens = sum(int(item["quantidade"] or 0) for item in itens)
    total_pagamentos = sum(float(pagamento["valor"] or 0) for pagamento in pagamentos)
    troco = max(total_pagamentos - float(venda["valor_total"] or 0), 0)
    auto_impressao = request.args.get("auto") == "1"

    conn.close()

    return render_template(
        "recibo_termico.html",
        venda=venda,
        itens=itens,
        pagamentos=pagamentos,
        config=config,
        total_itens=total_itens,
        quantidade_itens=quantidade_itens,
        troco=troco,
        auto_impressao=auto_impressao,
        retorno_url=url_for("vendas")
    )

@app.route("/vendas/<int:venda_id>/cancelar", methods=["POST"])
@admin_required
def cancelar_venda(venda_id):
    motivo_cancelamento = request.form.get("motivo_cancelamento", "").strip()

    if not motivo_cancelamento:
        flash("Informe o motivo do cancelamento.")
        return redirect(url_for("venda_detalhe", venda_id=venda_id))

    conn = get_db_connection()

    venda = conn.execute("""
        SELECT
            id,
            status,
            caixa_id,
            valor_total,
            forma_pagamento
        FROM vendas
        WHERE id = ?
    """, (venda_id,)).fetchone()

    if not venda:
        conn.close()
        flash("Venda não encontrada.")
        return redirect(url_for("vendas"))

    if venda["status"] == "CANCELADA":
        conn.close()
        flash("Esta venda já está cancelada.")
        return redirect(url_for("venda_detalhe", venda_id=venda_id))

    itens = conn.execute("""
        SELECT
            produto_id,
            quantidade
        FROM venda_itens
        WHERE venda_id = ?
    """, (venda_id,)).fetchall()

    for item in itens:
        produto = conn.execute("""
            SELECT
                id,
                estoque_atual
            FROM produtos
            WHERE id = ?
        """, (item["produto_id"],)).fetchone()

        if produto:
            estoque_anterior = produto["estoque_atual"]
            estoque_atualizado = estoque_anterior + item["quantidade"]

            conn.execute("""
                UPDATE produtos
                SET estoque_atual = ?
                WHERE id = ?
            """, (
                estoque_atualizado,
                item["produto_id"]
            ))

            conn.execute("""
                INSERT INTO movimentacoes_estoque (
                    produto_id,
                    tipo,
                    quantidade,
                    estoque_anterior,
                    estoque_atual,
                    observacoes
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                item["produto_id"],
                "CANCELAMENTO_VENDA",
                item["quantidade"],
                estoque_anterior,
                estoque_atualizado,
                f"Cancelamento da venda #{venda_id}"
            ))

    usuario = usuario_logado()
    cancelado_por = usuario["nome"] if usuario else "Sistema"

    conn.execute("""
        UPDATE vendas
        SET
            status = 'CANCELADA',
            motivo_cancelamento = ?,
            data_cancelamento = CURRENT_TIMESTAMP,
            cancelado_por = ?
        WHERE id = ?
    """, (
        motivo_cancelamento,
        cancelado_por,
        venda_id
    ))

    if venda["caixa_id"]:
        conn.execute("""
            INSERT INTO caixa_movimentacoes (
                caixa_id,
                tipo,
                valor,
                descricao,
                usuario_id,
                usuario_nome
            ) VALUES (?, ?, ?, ?, ?, ?)
        """, (
            venda["caixa_id"],
            "CANCELAMENTO_VENDA",
            venda["valor_total"] or 0,
            f"Cancelamento da venda #{venda_id} - {venda['forma_pagamento']}",
            usuario["id"],
            usuario["nome"]
        ))

    conn.commit()
    conn.close()

    registrar_log(
        "VENDA_CANCELADA",
        "vendas",
        venda_id,
        f"Venda #{venda_id} cancelada por {cancelado_por}. Motivo: {motivo_cancelamento}"
    )

    flash("Venda cancelada com sucesso. O estoque foi devolvido.")
    return redirect(url_for("venda_detalhe", venda_id=venda_id))

@app.route("/estoque", methods=["GET", "POST"])
@admin_required
def estoque():
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    if request.method == "POST":
        produto_id = request.form.get("produto_id")
        novo_estoque = int(request.form.get("novo_estoque") or 0)
        observacoes = request.form.get("observacoes")

        produto = conn.execute("""
            SELECT id, nome, estoque_atual
            FROM produtos
            WHERE id = ? AND ativo = 1
        """, (produto_id,)).fetchone()

        if not produto:
            flash("Produto não encontrado.")
            conn.close()
            return redirect(url_for("estoque"))

        if novo_estoque < 0:
            flash("O estoque não pode ser negativo.")
            conn.close()
            return redirect(url_for("estoque"))

        estoque_anterior = produto["estoque_atual"]
        diferenca = novo_estoque - estoque_anterior

        if diferenca > 0:
            tipo = "AJUSTE_ENTRADA"
            quantidade = diferenca
        elif diferenca < 0:
            tipo = "AJUSTE_SAIDA"
            quantidade = abs(diferenca)
        else:
            flash("Nenhuma alteração de estoque foi feita.")
            conn.close()
            return redirect(url_for("estoque"))

        cursor = conn.cursor()

        cursor.execute("""
            UPDATE produtos
            SET estoque_atual = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (
            novo_estoque,
            produto_id
        ))

        cursor.execute("""
            INSERT INTO movimentacoes_estoque (
                produto_id,
                tipo,
                quantidade,
                estoque_anterior,
                estoque_atual,
                observacoes
            ) VALUES (?, ?, ?, ?, ?, ?)
        """, (
            produto_id,
            tipo,
            quantidade,
            estoque_anterior,
            novo_estoque,
            observacoes or "Ajuste manual de estoque"
        ))

        conn.commit()
        conn.close()

        registrar_log(
            "ESTOQUE_AJUSTADO",
            "produtos",
            produto_id,
            f"Estoque ajustado de {estoque_anterior} para {novo_estoque}."
        )

        flash("Estoque ajustado com sucesso.")
        return redirect(url_for("estoque"))

    busca = request.args.get("busca", "").strip()
    status = request.args.get("status", "todos")

    page_produtos = int(request.args.get("page_produtos", 1))
    page_movimentacoes = normalizar_pagina(request.args.get("page_movimentacoes"))

    per_page_produtos = 10
    per_page_movimentacoes = 10

    if page_produtos < 1:
        page_produtos = 1

    if page_movimentacoes < 1:
        page_movimentacoes = 1

    offset_produtos = (page_produtos - 1) * per_page_produtos
    offset_movimentacoes = (page_movimentacoes - 1) * per_page_movimentacoes

    where_produtos = ["ativo = 1"]
    params_produtos = []

    if busca:
        termo = f"%{busca}%"
        where_produtos.append("""
            (
                nome LIKE ?
                OR categoria LIKE ?
                OR marca LIKE ?
                OR sku LIKE ?
            )
        """)
        params_produtos.extend([termo, termo, termo, termo])

    if status == "ok":
        where_produtos.append("estoque_atual > estoque_minimo")
    elif status == "baixo":
        where_produtos.append("estoque_atual > 0 AND estoque_atual <= estoque_minimo")
    elif status == "sem":
        where_produtos.append("estoque_atual = 0")

    where_produtos_sql = " AND ".join(where_produtos)

    total_produtos = conn.execute(f"""
        SELECT COUNT(*) AS total
        FROM produtos
        WHERE {where_produtos_sql}
    """, params_produtos).fetchone()["total"]

    total_paginas_produtos = (total_produtos + per_page_produtos - 1) // per_page_produtos

    if total_paginas_produtos == 0:
        total_paginas_produtos = 1

    if page_produtos > total_paginas_produtos:
        page_produtos = total_paginas_produtos
        offset_produtos = (page_produtos - 1) * per_page_produtos

    produtos = conn.execute(f"""
        SELECT
            id,
            nome,
            categoria,
            marca,
            sku,
            estoque_atual,
            estoque_minimo,
            preco_custo,
            preco_venda
        FROM produtos
        WHERE {where_produtos_sql}
        ORDER BY estoque_atual ASC, nome ASC
        LIMIT ? OFFSET ?
    """, params_produtos + [per_page_produtos, offset_produtos]).fetchall()

    total_movimentacoes = conn.execute("""
        SELECT COUNT(*) AS total
        FROM movimentacoes_estoque m
        INNER JOIN produtos p ON p.id = m.produto_id
    """).fetchone()["total"]

    page_movimentacoes, total_paginas_movimentacoes, offset_movimentacoes = calcular_paginacao(
        total_movimentacoes,
        page_movimentacoes,
        per_page_movimentacoes
    )

    movimentacoes = conn.execute("""
        SELECT
            m.id,
            m.tipo,
            m.quantidade,
            m.estoque_anterior,
            m.estoque_atual,
            m.observacoes,
            m.created_at,
            p.nome AS produto_nome,
            p.sku AS produto_sku
        FROM movimentacoes_estoque m
        INNER JOIN produtos p ON p.id = m.produto_id
        ORDER BY m.id DESC
        LIMIT ? OFFSET ?
    """, (
        per_page_movimentacoes,
        offset_movimentacoes
    )).fetchall()

    resumo_estoque = conn.execute("""
        SELECT
            COUNT(*) AS total_produtos,
            SUM(CASE WHEN estoque_atual = 0 THEN 1 ELSE 0 END) AS sem_estoque,
            SUM(CASE WHEN estoque_atual > 0 AND estoque_atual <= estoque_minimo THEN 1 ELSE 0 END) AS baixo_estoque,
            SUM(CASE WHEN estoque_atual > estoque_minimo THEN 1 ELSE 0 END) AS estoque_ok
        FROM produtos
        WHERE ativo = 1
    """).fetchone()

    produtos_ajuste = conn.execute("""
        SELECT id, nome, sku, estoque_atual
        FROM produtos
        WHERE ativo = 1
        ORDER BY nome ASC
    """).fetchall()

    conn.close()

    return render_template(
        "estoque.html",
        produtos=produtos,
        produtos_ajuste=produtos_ajuste,
        movimentacoes=movimentacoes,
        resumo_estoque=resumo_estoque,
        busca=busca,
        status=status,
        page_produtos=page_produtos,
        total_paginas_produtos=total_paginas_produtos,
        total_produtos=total_produtos,
        page_movimentacoes=page_movimentacoes,
        total_paginas_movimentacoes=total_paginas_movimentacoes,
        total_movimentacoes=total_movimentacoes,
        usuario=usuario_logado()
    )

@app.route("/relatorios")
@admin_required
def relatorios():
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")
    vendedor_id = request.args.get("vendedor_id", "").strip()
    cliente_id = request.args.get("cliente_id", "").strip()
    forma_pagamento = request.args.get("forma_pagamento", "").strip()

    hoje = date.today().isoformat()

    if not data_inicio:
        data_inicio = hoje

    if not data_fim:
        data_fim = hoje

    try:
        data_inicio_obj = date.fromisoformat(data_inicio)
        data_fim_obj = date.fromisoformat(data_fim)
    except (TypeError, ValueError):
        data_inicio_obj = date.today()
        data_fim_obj = date.today()

    if data_inicio_obj > data_fim_obj:
        data_inicio_obj, data_fim_obj = data_fim_obj, data_inicio_obj

    data_inicio = data_inicio_obj.isoformat()
    data_fim = data_fim_obj.isoformat()

    quantidade_dias_periodo = (data_fim_obj - data_inicio_obj).days + 1
    data_fim_anterior_obj = data_inicio_obj - timedelta(days=1)
    data_inicio_anterior_obj = data_fim_anterior_obj - timedelta(days=quantidade_dias_periodo - 1)

    atalhos_periodo = [
        {
            "label": "Hoje",
            "inicio": date.today().isoformat(),
            "fim": date.today().isoformat(),
        },
        {
            "label": "7 dias",
            "inicio": (date.today() - timedelta(days=6)).isoformat(),
            "fim": date.today().isoformat(),
        },
        {
            "label": "30 dias",
            "inicio": (date.today() - timedelta(days=29)).isoformat(),
            "fim": date.today().isoformat(),
        },
        {
            "label": "Este mês",
            "inicio": date.today().replace(day=1).isoformat(),
            "fim": date.today().isoformat(),
        },
    ]

    periodo_anterior_label = (
        f"{data_inicio_anterior_obj.strftime('%d/%m/%Y')} a "
        f"{data_fim_anterior_obj.strftime('%d/%m/%Y')}"
    )

    page = int(request.args.get("page", 1))
    per_page = 10

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    where_clauses = [
        "date(v.data_venda) BETWEEN date(?) AND date(?)",
        "v.status = 'CONCLUIDA'"
    ]

    params = [data_inicio, data_fim]

    if vendedor_id:
        where_clauses.append("v.vendedor_id = ?")
        params.append(vendedor_id)

    if cliente_id:
        where_clauses.append("v.cliente_id = ?")
        params.append(cliente_id)

    if forma_pagamento:
        where_clauses.append("""
            EXISTS (
                SELECT 1
                FROM venda_pagamentos vp_filtro
                WHERE vp_filtro.venda_id = v.id
                  AND vp_filtro.forma_pagamento = ?
            )
        """)
        params.append(forma_pagamento)

    where_sql = " AND ".join(where_clauses)

    conn = get_db_connection()

    total_registros = conn.execute(f"""
        SELECT COUNT(*) AS total
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE {where_sql}
    """, params).fetchone()["total"]

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas == 0:
        total_paginas = 1

    if page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    vendas = conn.execute(f"""
        SELECT
            v.id,
            v.data_venda,
            COALESCE(c.nome, 'Cliente não identificado') AS cliente_nome,
            v.vendedor,
            v.forma_pagamento,
            v.valor_total,
            v.custo_total,
            v.lucro_total,
            v.desconto_total,
            v.status,
            (
                SELECT GROUP_CONCAT(pagamento.forma_pagamento, ' + ')
                FROM venda_pagamentos pagamento
                WHERE pagamento.venda_id = v.id
            ) AS pagamentos_descricao
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE {where_sql}
        ORDER BY v.data_venda DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()

    resumo = conn.execute(f"""
        SELECT
            COUNT(*) AS quantidade_vendas,
            COALESCE(SUM(v.valor_total), 0) AS total_vendido,
            COALESCE(SUM(v.custo_total), 0) AS total_custo,
            COALESCE(SUM(v.lucro_total), 0) AS total_lucro,
            COALESCE(SUM(v.desconto_total), 0) AS total_desconto
        FROM vendas v
        WHERE {where_sql}
    """, params).fetchone()

    params_periodo_anterior = [
        data_inicio_anterior_obj.isoformat(),
        data_fim_anterior_obj.isoformat(),
        *params[2:],
    ]

    resumo_periodo_anterior = conn.execute(f"""
        SELECT
            COUNT(*) AS quantidade_vendas,
            COALESCE(SUM(v.valor_total), 0) AS total_vendido,
            COALESCE(SUM(v.custo_total), 0) AS total_custo,
            COALESCE(SUM(v.lucro_total), 0) AS total_lucro,
            COALESCE(SUM(v.desconto_total), 0) AS total_desconto
        FROM vendas v
        WHERE {where_sql}
    """, params_periodo_anterior).fetchone()

    serie_diaria_rows = conn.execute(f"""
        SELECT
            date(v.data_venda) AS dia,
            COALESCE(SUM(v.valor_total), 0) AS total_vendido,
            COALESCE(SUM(v.lucro_total), 0) AS total_lucro
        FROM vendas v
        WHERE {where_sql}
        GROUP BY date(v.data_venda)
        ORDER BY date(v.data_venda)
    """, params).fetchall()

    formas_pagamento_rows = conn.execute(f"""
        SELECT
            vp.forma_pagamento,
            COUNT(DISTINCT vp.venda_id) AS quantidade,
            COALESCE(SUM(vp.valor), 0) AS total
        FROM venda_pagamentos vp
        INNER JOIN vendas v ON v.id = vp.venda_id
        WHERE {where_sql}
        GROUP BY vp.forma_pagamento
        ORDER BY total DESC
    """, params).fetchall()

    produtos_mais_vendidos = conn.execute(f"""
        SELECT
            p.nome AS produto_nome,
            p.sku,
            SUM(vi.quantidade) AS quantidade_total,
            SUM(vi.subtotal) AS total_vendido
        FROM venda_itens vi
        INNER JOIN produtos p ON p.id = vi.produto_id
        INNER JOIN vendas v ON v.id = vi.venda_id
        WHERE {where_sql}
        GROUP BY p.id
        ORDER BY quantidade_total DESC
        LIMIT 10
    """, params).fetchall()

    clientes_mais_compraram = conn.execute(f"""
        SELECT
            c.nome AS cliente_nome,
            c.telefone,
            COUNT(v.id) AS quantidade_compras,
            SUM(v.valor_total) AS total_comprado
        FROM vendas v
        INNER JOIN clientes c ON c.id = v.cliente_id
        WHERE {where_sql}
        GROUP BY c.id
        ORDER BY total_comprado DESC
        LIMIT 10
    """, params).fetchall()

    where_canceladas_clauses = [
        "date(COALESCE(v.data_cancelamento, v.data_venda)) BETWEEN date(?) AND date(?)",
        "v.status = 'CANCELADA'"
    ]

    params_canceladas = [data_inicio, data_fim]

    if vendedor_id:
        where_canceladas_clauses.append("v.vendedor_id = ?")
        params_canceladas.append(vendedor_id)

    if cliente_id:
        where_canceladas_clauses.append("v.cliente_id = ?")
        params_canceladas.append(cliente_id)

    if forma_pagamento:
        where_canceladas_clauses.append("""
            EXISTS (
                SELECT 1
                FROM venda_pagamentos vp_filtro
                WHERE vp_filtro.venda_id = v.id
                  AND vp_filtro.forma_pagamento = ?
            )
        """)
        params_canceladas.append(forma_pagamento)

    where_canceladas_sql = " AND ".join(where_canceladas_clauses)

    resumo_canceladas = conn.execute(f"""
        SELECT
            COUNT(*) AS quantidade_canceladas,
            COALESCE(SUM(v.valor_total), 0) AS total_cancelado,
            COALESCE(SUM(v.lucro_total), 0) AS lucro_cancelado
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE {where_canceladas_sql}
    """, params_canceladas).fetchone()

    vendas_canceladas = conn.execute(f"""
        SELECT
            v.id,
            v.data_venda,
            v.data_cancelamento,
            v.cancelado_por,
            v.motivo_cancelamento,
            COALESCE(c.nome, 'Cliente não identificado') AS cliente_nome,
            v.vendedor,
            v.forma_pagamento,
            v.valor_total,
            v.lucro_total
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE {where_canceladas_sql}
        ORDER BY COALESCE(v.data_cancelamento, v.data_venda) DESC
        LIMIT 10
    """, params_canceladas).fetchall()

    vendedores = conn.execute("""
        SELECT id, nome
        FROM usuarios
        WHERE perfil = 'VENDEDOR'
          AND ativo = 1
        ORDER BY nome ASC
    """).fetchall()

    clientes = conn.execute("""
        SELECT id, nome, telefone
        FROM clientes
        WHERE ativo = 1
        ORDER BY nome ASC
    """).fetchall()

    formas_pagamento_opcoes = conn.execute("""
        SELECT DISTINCT forma_pagamento
        FROM venda_pagamentos
        WHERE forma_pagamento IS NOT NULL
          AND forma_pagamento != ''
        ORDER BY forma_pagamento ASC
    """).fetchall()

    quantidade_vendas = resumo["quantidade_vendas"] or 0
    total_vendido = resumo["total_vendido"] or 0
    total_lucro = resumo["total_lucro"] or 0

    ticket_medio = total_vendido / quantidade_vendas if quantidade_vendas > 0 else 0
    margem_media = (total_lucro / total_vendido * 100) if total_vendido > 0 else 0

    quantidade_vendas_anterior = resumo_periodo_anterior["quantidade_vendas"] or 0
    total_vendido_anterior = resumo_periodo_anterior["total_vendido"] or 0
    total_lucro_anterior = resumo_periodo_anterior["total_lucro"] or 0
    ticket_medio_anterior = (
        total_vendido_anterior / quantidade_vendas_anterior
        if quantidade_vendas_anterior > 0
        else 0
    )
    margem_media_anterior = (
        total_lucro_anterior / total_vendido_anterior * 100
        if total_vendido_anterior > 0
        else 0
    )

    def calcular_variacao_percentual(valor_atual, valor_anterior):
        valor_atual = float(valor_atual or 0)
        valor_anterior = float(valor_anterior or 0)

        if abs(valor_anterior) <= 0.009:
            return None if abs(valor_atual) <= 0.009 else 100.0

        return round(((valor_atual - valor_anterior) / abs(valor_anterior)) * 100, 1)

    variacoes = {
        "quantidade_vendas": calcular_variacao_percentual(
            quantidade_vendas,
            quantidade_vendas_anterior,
        ),
        "total_vendido": calcular_variacao_percentual(
            total_vendido,
            total_vendido_anterior,
        ),
        "total_lucro": calcular_variacao_percentual(
            total_lucro,
            total_lucro_anterior,
        ),
        "ticket_medio": calcular_variacao_percentual(
            ticket_medio,
            ticket_medio_anterior,
        ),
        "margem_media": round(margem_media - margem_media_anterior, 1),
    }

    dados_diarios = {
        date.fromisoformat(row["dia"]): {
            "total_vendido": float(row["total_vendido"] or 0),
            "total_lucro": float(row["total_lucro"] or 0),
        }
        for row in serie_diaria_rows
        if row["dia"]
    }

    limite_colunas_grafico = 16
    tamanho_bloco = max(
        1,
        (quantidade_dias_periodo + limite_colunas_grafico - 1)
        // limite_colunas_grafico,
    )
    serie_temporal = []

    for deslocamento in range(0, quantidade_dias_periodo, tamanho_bloco):
        inicio_bloco = data_inicio_obj + timedelta(days=deslocamento)
        fim_bloco = min(
            inicio_bloco + timedelta(days=tamanho_bloco - 1),
            data_fim_obj,
        )
        cursor_dia = inicio_bloco
        total_bloco = 0
        lucro_bloco = 0

        while cursor_dia <= fim_bloco:
            valores_dia = dados_diarios.get(cursor_dia, {})
            total_bloco += valores_dia.get("total_vendido", 0)
            lucro_bloco += valores_dia.get("total_lucro", 0)
            cursor_dia += timedelta(days=1)

        if inicio_bloco == fim_bloco:
            label_bloco = inicio_bloco.strftime("%d/%m")
        else:
            label_bloco = (
                f"{inicio_bloco.strftime('%d/%m')}–"
                f"{fim_bloco.strftime('%d/%m')}"
            )

        serie_temporal.append({
            "label": label_bloco,
            "total_vendido": total_bloco,
            "total_lucro": lucro_bloco,
        })

    maior_valor_serie = max(
        [
            max(item["total_vendido"], item["total_lucro"], 0)
            for item in serie_temporal
        ]
        or [0]
    )

    for item in serie_temporal:
        if maior_valor_serie > 0:
            percentual_total = item["total_vendido"] / maior_valor_serie * 100
            percentual_lucro = max(item["total_lucro"], 0) / maior_valor_serie * 100
            item["altura_total"] = max(4, round(percentual_total, 2)) if item["total_vendido"] > 0 else 0
            item["altura_lucro"] = max(4, round(percentual_lucro, 2)) if item["total_lucro"] > 0 else 0
        else:
            item["altura_total"] = 0
            item["altura_lucro"] = 0

    total_formas_pagamento = sum(
        float(item["total"] or 0)
        for item in formas_pagamento_rows
    )
    formas_pagamento = []

    for item in formas_pagamento_rows:
        valor_forma = float(item["total"] or 0)
        formas_pagamento.append({
            "forma_pagamento": item["forma_pagamento"],
            "quantidade": item["quantidade"],
            "total": valor_forma,
            "percentual": (
                round(valor_forma / total_formas_pagamento * 100, 1)
                if total_formas_pagamento > 0
                else 0
            ),
        })

    maior_total_produto = max(
        [float(item["total_vendido"] or 0) for item in produtos_mais_vendidos]
        or [0]
    )
    produtos_ranking = [
        {
            **dict(item),
            "percentual": (
                round(float(item["total_vendido"] or 0) / maior_total_produto * 100, 1)
                if maior_total_produto > 0
                else 0
            ),
        }
        for item in produtos_mais_vendidos
    ]

    maior_total_cliente = max(
        [float(item["total_comprado"] or 0) for item in clientes_mais_compraram]
        or [0]
    )
    clientes_ranking = [
        {
            **dict(item),
            "percentual": (
                round(float(item["total_comprado"] or 0) / maior_total_cliente * 100, 1)
                if maior_total_cliente > 0
                else 0
            ),
        }
        for item in clientes_mais_compraram
    ]

    periodo_label = (
        data_inicio_obj.strftime("%d/%m/%Y")
        if data_inicio_obj == data_fim_obj
        else (
            f"{data_inicio_obj.strftime('%d/%m/%Y')} a "
            f"{data_fim_obj.strftime('%d/%m/%Y')}"
        )
    )
    filtros_ativos = sum(bool(valor) for valor in (
        vendedor_id,
        cliente_id,
        forma_pagamento,
    ))

    conn.close()

    return render_template(
        "relatorios.html",
        vendas=vendas,
        resumo=resumo,
        formas_pagamento=formas_pagamento,
        produtos_mais_vendidos=produtos_mais_vendidos,
        clientes_mais_compraram=clientes_mais_compraram,
        resumo_canceladas=resumo_canceladas,
        vendas_canceladas=vendas_canceladas,
        vendedores=vendedores,
        clientes=clientes,
        formas_pagamento_opcoes=formas_pagamento_opcoes,
        data_inicio=data_inicio,
        data_fim=data_fim,
        vendedor_id=vendedor_id,
        cliente_id=cliente_id,
        forma_pagamento=forma_pagamento,
        ticket_medio=ticket_medio,
        margem_media=margem_media,
        variacoes=variacoes,
        serie_temporal=serie_temporal,
        produtos_ranking=produtos_ranking,
        clientes_ranking=clientes_ranking,
        atalhos_periodo=atalhos_periodo,
        periodo_label=periodo_label,
        periodo_anterior_label=periodo_anterior_label,
        filtros_ativos=filtros_ativos,
        page=page,
        total_paginas=total_paginas,
        total_registros=total_registros,
        usuario=usuario_logado()
    )

@app.route("/relatorios/caixas")
@admin_required
def relatorio_caixas():
    conn = get_db_connection()

    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip()
    usuario_id = request.args.get("usuario_id", "").strip()
    status = request.args.get("status", "").strip()
    somente_diferenca = request.args.get("somente_diferenca", "").strip()

    if not data_inicio:
        data_inicio = datetime.now().strftime("%Y-%m-01")

    if not data_fim:
        data_fim = datetime.now().strftime("%Y-%m-%d")

    filtros = [
        "date(c.data_abertura) BETWEEN date(?) AND date(?)"
    ]

    params = [data_inicio, data_fim]

    if usuario_id:
        filtros.append("c.usuario_abertura_id = ?")
        params.append(usuario_id)

    if status:
        filtros.append("c.status = ?")
        params.append(status)

    if somente_diferenca == "1":
        filtros.append("ABS(COALESCE(c.diferenca, 0)) > 0.009")

    where_sql = " AND ".join(filtros)

    caixas = conn.execute(f"""
        SELECT
            c.id,
            c.data_abertura,
            c.data_fechamento,
            c.usuario_abertura_nome,
            c.usuario_fechamento_nome,
            c.valor_inicial,
            c.total_vendas,
            c.total_dinheiro,
            c.total_pix,
            c.total_cartao,
            c.total_outros,
            c.entradas_manuais,
            c.saidas_manuais,
            c.valor_esperado,
            c.valor_informado,
            c.diferenca,
            c.status
        FROM caixas c
        WHERE {where_sql}
        ORDER BY c.id DESC
    """, params).fetchall()

    resumo = conn.execute(f"""
        SELECT
            COUNT(*) AS quantidade_caixas,
            COALESCE(SUM(CASE WHEN c.status = 'ABERTO' THEN 1 ELSE 0 END), 0) AS caixas_abertos,
            COALESCE(SUM(CASE WHEN c.status = 'FECHADO' THEN 1 ELSE 0 END), 0) AS caixas_fechados,
            COALESCE(SUM(
                CASE WHEN ABS(COALESCE(c.diferenca, 0)) > 0.009 THEN 1 ELSE 0 END
            ), 0) AS caixas_com_diferenca,
            COALESCE(SUM(c.valor_inicial), 0) AS total_inicial,
            COALESCE(SUM(c.total_vendas), 0) AS total_vendas,
            COALESCE(SUM(c.total_dinheiro), 0) AS total_dinheiro,
            COALESCE(SUM(c.total_pix), 0) AS total_pix,
            COALESCE(SUM(c.total_cartao), 0) AS total_cartao,
            COALESCE(SUM(c.total_outros), 0) AS total_outros,
            COALESCE(SUM(c.entradas_manuais), 0) AS total_entradas,
            COALESCE(SUM(c.saidas_manuais), 0) AS total_saidas,
            COALESCE(SUM(c.valor_esperado), 0) AS total_esperado,
            COALESCE(SUM(c.valor_informado), 0) AS total_informado,
            COALESCE(SUM(c.diferenca), 0) AS total_diferenca,
            COALESCE(SUM(
                CASE WHEN COALESCE(c.diferenca, 0) > 0.009 THEN c.diferenca ELSE 0 END
            ), 0) AS total_sobras,
            COALESCE(SUM(
                CASE WHEN COALESCE(c.diferenca, 0) < -0.009 THEN ABS(c.diferenca) ELSE 0 END
            ), 0) AS total_faltas
        FROM caixas c
        WHERE {where_sql}
    """, params).fetchone()

    usuarios = conn.execute("""
        SELECT
            id,
            nome
        FROM usuarios
        WHERE ativo = 1
        ORDER BY nome
    """).fetchall()

    conn.close()

    return render_template(
        "relatorio_caixas.html",
        caixas=caixas,
        resumo=resumo,
        usuarios=usuarios,
        data_inicio=data_inicio,
        data_fim=data_fim,
        usuario_id=usuario_id,
        status=status,
        somente_diferenca=somente_diferenca,
        usuario=usuario_logado()
    )

@app.route("/relatorios/exportar")
@admin_required
def exportar_relatorio():
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")
    vendedor_id = request.args.get("vendedor_id", "").strip()
    cliente_id = request.args.get("cliente_id", "").strip()
    forma_pagamento = request.args.get("forma_pagamento", "").strip()

    hoje = date.today().isoformat()

    if not data_inicio:
        data_inicio = hoje

    if not data_fim:
        data_fim = hoje

    where_clauses = [
        "date(v.data_venda) BETWEEN date(?) AND date(?)",
        "v.status = 'CONCLUIDA'"
    ]

    params = [data_inicio, data_fim]

    if vendedor_id:
        where_clauses.append("v.vendedor_id = ?")
        params.append(vendedor_id)

    if cliente_id:
        where_clauses.append("v.cliente_id = ?")
        params.append(cliente_id)

    if forma_pagamento:
        where_clauses.append("""
            EXISTS (
                SELECT 1
                FROM venda_pagamentos vp_filtro
                WHERE vp_filtro.venda_id = v.id
                  AND vp_filtro.forma_pagamento = ?
            )
        """)
        params.append(forma_pagamento)

    where_sql = " AND ".join(where_clauses)

    conn = get_db_connection()

    vendas = conn.execute(f"""
        SELECT
            v.id,
            v.data_venda,
            c.nome AS cliente_nome,
            COALESCE(c.telefone, '-') AS cliente_telefone,
            v.vendedor,
            v.forma_pagamento,
            v.desconto_total,
            v.valor_total,
            v.custo_total,
            v.lucro_total,
            v.status,
            (
                SELECT GROUP_CONCAT(pagamento.forma_pagamento, ' + ')
                FROM venda_pagamentos pagamento
                WHERE pagamento.venda_id = v.id
            ) AS pagamentos_descricao
        FROM vendas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE {where_sql}
        ORDER BY v.data_venda DESC
    """, params).fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"

    ws.append([
        "ID Venda",
        "Data",
        "Cliente",
        "Telefone",
        "Vendedor",
        "Forma de pagamento",
        "Desconto",
        "Valor total",
        "Custo total",
        "Lucro total",
        "Status"
    ])

    for venda in vendas:
        ws.append([
            venda["id"],
            venda["data_venda"],
            venda["cliente_nome"],
            venda["cliente_telefone"],
            venda["vendedor"],
            venda["pagamentos_descricao"] or venda["forma_pagamento"],
            venda["desconto_total"],
            venda["valor_total"],
            venda["custo_total"],
            venda["lucro_total"],
            venda["status"]
        ])

    nome_arquivo = f"relatorio_vendas_{data_inicio}_a_{data_fim}.xlsx"
    diretorio_exportacoes = os.path.join(BASE_DIR, "exports")
    os.makedirs(diretorio_exportacoes, exist_ok=True)
    caminho_arquivo = os.path.join(diretorio_exportacoes, nome_arquivo)

    wb.save(caminho_arquivo)

    return send_file(caminho_arquivo, as_attachment=True)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
